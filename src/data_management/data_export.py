import sqlite3
import sys
import os
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv

# Add project root to path for imports
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from src.data_management.data_import import get_site_id_from_domain_name
from src.core.filters import check_for_node, is_high_priority
from openpyxl.worksheet.datavalidation import DataValidation

def get_all_sites():


    with open("sql/get_all_sites.sql", 'r') as file:
        sql_query = file.read()
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
        cursor.execute(sql_query)
        results = cursor.fetchall()
        results = [result[0] for result in results]

        if not results:
            return []
        return results



def get_pdf_reports_by_site_name(site_name):

    with open("sql/get_pdf_reports_by_site_name.sql", 'r') as file:
        sql_query = file.read()
        formatted_query = sql_query.format(site_name=site_name)

        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()

        # Execute the SQL query
        cursor.execute(formatted_query)
        results = cursor.fetchall()

        # If there are no results, return an empty list
        if not results:
            return []

        # Get column names from the cursor
        col_names = [desc[0] for desc in cursor.description]
        # Create a namedtuple class with column names
        Row = namedtuple('Row', col_names)

        # Convert sqlite3.Row objects to namedtuples
        results = [Row(*row) for row in results]

        # Close the database connection
        conn.close()

        return results


def get_pdfs_by_site_name(site_name):

    with open("sql/get_pdfs_by_domain_name.sql", 'r') as file:
        sql_query = file.read()
        formatted_query = sql_query.format(site_name=site_name)
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
        cursor.execute(formatted_query)
        results = cursor.fetchall()
        return results


#
# print(get_pdfs_by_site_name('creativewriting.sfsu.edu'))



def get_all_users_with_pdfs():
    with open("sql/get_all_users_with_pdf_files.sql", 'r') as file:
        sql_query = file.read()
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
        cursor.execute(sql_query)
        results = cursor.fetchall()
        results = [result for result in results]

        if not results:
            return []
        return results


def get_site_failures(site_name):

    with open("sql/get_failures_by_site_id.sql", 'r') as file:
        sql_query = file.read()

        # need to get site id from site name

        site_id = get_site_id_from_domain_name(site_name.replace("-", "."))
        formatted_query = sql_query.format(site_id=site_id)
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()

        cursor.execute(formatted_query)
        results = cursor.fetchall()

        # If there are no results, return an empty list
        if not results:
            return []


        # Get column names from the cursor
        col_names = [desc[0] for desc in cursor.description]
        # Create a namedtuple class with column names
        Row = namedtuple('Row', col_names)

        # Convert sqlite3.Row objects to namedtuples
        results = [Row(*row) for row in results]

        # Close the database connection
        conn.close()

        return results



def write_data_to_excel(data, failure_data, file_name="output.xlsx"):
    # Create a workbook
    wb = Workbook()

    # Create "Scanned PDFs" worksheet and set it as the active worksheet
    scanned_pdfs_ws = wb.create_sheet("Scanned PDFs", 0)
    wb.active = 0

    # Create "Unique PDFs" worksheet
    unique_pdfs_ws = wb.create_sheet("Unique PDFs", 1)

    # Create "Failure" worksheet
    failure_ws = wb.create_sheet("Failure", 2)

    # Create "Instructions" worksheet
    instructions_ws = wb.create_sheet("Instructions", 3)

    def add_data_to_scanned_pdfs(data, worksheet):
        if not data:
            print("No data to write to Excel.")
            return

        def pdf_title_from_url(url: str) -> str:
            try:
                from urllib.parse import urlsplit, unquote

                path = urlsplit(str(url)).path
                name = path.rsplit('/', 1)[-1] if path else str(url)
                name = unquote(name)
                if name.lower().endswith('.pdf'):
                    name = name[:-4]
                name = name.replace('_', ' ').replace('-', ' ').strip()
                return ' '.join(name.split()) if name else str(url)
            except Exception:
                return str(url)

        # 1) Build the columns list from the namedtuple fields
        columns = list(data[0]._fields)
        # Rename 'file_hash' to 'fingerprint'
        columns = ["fingerprint" if col == "file_hash" else col for col in columns]
        # Remove the 'box_folder' column
        columns.remove('box_folder')

        # Add a friendly title column first (derived from pdf_uri)
        columns.insert(0, 'pdf_title')

        # Add custom columns to the end
        columns.append("Errors/Page")
        columns.append("Low Priority")
        columns.append("DPRC Will Remediate")

        # 2) Write the column headers to the worksheet
        worksheet.append(columns)

        # 3) Define the fill color for high priority rows
        red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

        # 4) Create data validations for "Low Priority" and "DPRC Will Remediate" columns
        dv_low_priority = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        dv_dprc_remediate = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)

        # Add both data validations to the worksheet
        worksheet.add_data_validation(dv_low_priority)
        worksheet.add_data_validation(dv_dprc_remediate)

        # Loop through data rows and populate cells
        for item in data:
            item_list = list(item)

            high_priority = is_high_priority(item)  # (Your function to check if row is high priority)

            # Skip if the second item is a node link
            if not check_for_node(item_list[1]):    # (Your function to check if node link)
                # Convert first two columns to hyperlinks
                item_list[0] = f'=HYPERLINK("{item[0]}", "{item[0]}")'
                item_list[1] = f'=HYPERLINK("{item[1]}", "{item[1].split(" ")[0]}")' # remove accidental datatime info aadded

                # Truncate the 4th item (index 3) to 6 characters
                item_list[3] = item[3][0:6]

                # Convert these columns to "Yes"/"No"
                item_list[8] = "Yes" if item_list[8] == 1 else "No"
                item_list[10] = "Yes" if item_list[10] == 1 else "No"
                item_list[11] = "Yes" if item_list[11] == 1 else "No"
                item_list[13] = "Yes" if item_list[13] == 1 else "No"

                # Remove the box.com link at index 14
                del item_list[14]

                # Calculate Errors/Page
                errors_per_page = 0
                if item[7] != 0 and item[12] != 0:
                    errors_per_page = round(int(item[7]) / int(item[12]))
                item_list.append(errors_per_page)

                # Append placeholders for Low Priority & DPRC columns
                item_list.append("")
                item_list.append("")

                # Insert the PDF title as the first column (keeps existing index-based logic intact above)
                item_list.insert(0, pdf_title_from_url(item[0]))

                # Append the row to the worksheet
                worksheet.append(item_list)
                current_row = worksheet._current_row

                # Identify columns for data validation
                low_priority_col_idx = len(columns) - 1
                dprc_remediate_col_idx = len(columns)

                # Set Low Priority based on high_priority flag
                # Low Priority = Yes means NOT high priority (safe to defer)
                # Low Priority = No means IS high priority (needs immediate attention)
                low_priority_cell = worksheet.cell(row=current_row, column=low_priority_col_idx)
                dprc_remediate_cell = worksheet.cell(row=current_row, column=dprc_remediate_col_idx)
                low_priority_cell.value = "No" if high_priority else "Yes"
                dprc_remediate_cell.value = "No"

                dv_low_priority.add(low_priority_cell)
                dv_dprc_remediate.add(dprc_remediate_cell)

                # If high_priority is True, apply red_fill to the entire row
                if high_priority:
                    for cell in worksheet[current_row]:
                        cell.fill = red_fill

        # Make PDF URL + parent URL appear as hyperlinks (blue + underline)
        max_row = worksheet.max_row
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=3, max_row=max_row):
            for cell in row:
                cell.font = Font(color='0563C1', underline='single')

        # Convert the data into a formatted Table
        table_range = f"A1:{chr(64 + len(columns))}{max_row}"
        data_table = Table(displayName="DataTable", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        data_table.tableStyleInfo = style
        worksheet.add_table(data_table)

    def add_data_to_failure(failure_data, worksheet):
        if not failure_data:
            print("No failure data to write to Excel.")
            return

        # Collect columns from the first item
        columns = list(failure_data[0]._fields)
        worksheet.append(columns)

        for item in failure_data:
            worksheet.append(list(item))

        # Create a table in the Failure sheet
        table_range = f"A1:{chr(64 + len(columns))}{len(failure_data) + 1}"
        failure_table = Table(displayName="FailureDataTable", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        failure_table.tableStyleInfo = style
        worksheet.add_table(failure_table)

    def add_data_to_unique_pdfs(data, worksheet):
        if not data:
            print("No data to write to Excel.")
            return

        def pdf_title_from_url(url: str) -> str:
            try:
                from urllib.parse import urlsplit, unquote

                path = urlsplit(str(url)).path
                name = path.rsplit('/', 1)[-1] if path else str(url)
                name = unquote(name)
                if name.lower().endswith('.pdf'):
                    name = name[:-4]
                name = name.replace('_', ' ').replace('-', ' ').strip()
                return ' '.join(name.split()) if name else str(url)
            except Exception:
                return str(url)

        columns = [
            "pdf_title",
            "pdf_uri",
            "fingerprint",
            "domain_name",
            "violations",
            "failed_checks",
            "tagged",
            "pdf_text_type",
            "title_set",
            "language_set",
            "page_count",
            "has_form",
            "approved_pdf_exporter",
            "link_count",
            "sample_parent_uri",
        ]
        worksheet.append(columns)

        # Build an aggregation by pdf_uri
        by_pdf = {}
        for item in data:
            item_list = list(item)
            # Skip if parent_uri is a node link
            if check_for_node(item_list[1]):
                continue
            pdf_uri = getattr(item, "pdf_uri", None)
            parent_uri = getattr(item, "parent_uri", None)
            if not pdf_uri:
                continue

            if pdf_uri not in by_pdf:
                by_pdf[pdf_uri] = {
                    "item": item,
                    "parents": set(),
                }
            if parent_uri:
                by_pdf[pdf_uri]["parents"].add(str(parent_uri).split(" ")[0])

        for pdf_uri, agg in by_pdf.items():
            item = agg["item"]
            parents = sorted(agg["parents"]) if agg["parents"] else []
            sample_parent = parents[0] if parents else ""
            file_hash = getattr(item, "file_hash", "")
            fingerprint = (file_hash or "")[0:6]

            row = [
                pdf_title_from_url(pdf_uri),
                f'=HYPERLINK("{pdf_uri}", "{pdf_uri}")',
                fingerprint,
                getattr(item, "domain_name", ""),
                getattr(item, "violations", ""),
                getattr(item, "failed_checks", ""),
                "Yes" if getattr(item, "tagged", 0) == 1 else "No",
                getattr(item, "pdf_text_type", ""),
                getattr(item, "title_set", ""),
                getattr(item, "language_set", ""),
                getattr(item, "page_count", ""),
                "Yes" if getattr(item, "has_form", 0) == 1 else "No",
                "Yes" if getattr(item, "approved_pdf_exporter", 0) == 1 else "No",
                len(parents),
                f'=HYPERLINK("{sample_parent}", "{sample_parent}")' if sample_parent else "",
            ]
            worksheet.append(row)

        # Make hyperlink columns appear as hyperlinks (blue + underline)
        max_row = worksheet.max_row
        # pdf_uri is now column B
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=max_row):
            for cell in row:
                cell.font = Font(color='0563C1', underline='single')
        for row in worksheet.iter_rows(min_row=2, min_col=len(columns), max_col=len(columns), max_row=max_row):
            for cell in row:
                if cell.value:
                    cell.font = Font(color='0563C1', underline='single')

        # Convert the data into a formatted Table
        table_range = f"A1:{chr(64 + len(columns))}{max_row}"
        data_table = Table(displayName="UniquePdfTable", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        data_table.tableStyleInfo = style
        worksheet.add_table(data_table)

    def add_instructions(worksheet):
        cell = worksheet.cell(row=1, column=1)
        cell.value = (
            "Important Instructions:\n\n"
            "1) Please review PDF accessibility requirements at: https://www.calstatela.edu/accessibility\n"
            "2) Cal State LA will provide access to PDF remediation tools for all employees.\n"
            "3) Please update the Priority column if a PDF meets the requirements for low priority. Only focus on RED highlights.\n"
            "4) Pay Attention to the 'fingerprint' column. This is the unique identifier for each PDF. There may be duplicates on your domain.\n"
            "5) If you need assistance with PDF remediation, please contact:\n"
            "   Email: accessibility@calstatela.edu | Phone: 323-343-6170 (ITS Help Desk)\n"
            "6) These scans only look at files hosted on drupal or box, if you feel files are missing please let us know and we can run a new scan.\n"
            "7) For questions, training, or feedback, contact: accessibility@calstatela.edu\n"
        )
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        cell.font = Font(bold=True)
        worksheet.column_dimensions['A'].width = 120
        worksheet.row_dimensions[1].height = 300

    # ---------------- MAIN LOGIC ----------------

    # 1) Populate the "Scanned PDFs" sheet
    add_data_to_scanned_pdfs(data, scanned_pdfs_ws)

    # 1b) Populate the "Unique PDFs" sheet
    add_data_to_unique_pdfs(data, unique_pdfs_ws)

    # 2) Create the "Failure" sheet
    add_data_to_failure(failure_data, failure_ws)

    # 3) Add instructions on a dedicated worksheet (keeps Scanned PDFs table clean)
    add_instructions(instructions_ws)

    # 4) Save the workbook
    wb.save(file_name)
    print(f"Data written to {file_name} with a table format.")



