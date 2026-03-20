"""Single source of truth for all Excel reading, parsing, and shared utilities.

Every reporting script (historical_analysis, generate_master_report,
generate_master_report_html, sharepoint_sync) imports ALL shared logic
from here. No script owns its own version of any function that another
script also uses — that was the root cause of count mismatches.

Sections
--------
1. Type coercion utilities        (_coerce_int, _coerce_bool)
2. Excel cell utilities           (_strip_hyperlink)
3. Filename / path utilities      (_parse_timestamp, find_latest_xlsx,
                                   folder_to_display_name, xlsx_report_date)
4. Excel row reader               (read_pdf_rows)
5. Row-level priority helper      (row_to_priority_data)
6. Aggregate Excel parser         (parse_domain_excel)  ← counts are computed here
7. Domain scan collector          (collect_from_local)
8. HTML / template utilities      (_js)

Priority source of truth
------------------------
The 'Low Priority' column in each Excel file is set by data_export.py at scan
time using is_high_priority() from src/core/filters.py.  All scripts read that
stamped value — never re-evaluating the filter logic from raw row data — so every
report shows identical counts for any given run.
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import config
import openpyxl

# Timestamp pattern embedded in report filenames:
# e.g.  calstatela-edu_ecst-2026-03-20_08-14-32.xlsx
_TIMESTAMP_RE = re.compile(
    r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE
)

# Hyperlink formula pattern used by Excel HYPERLINK() cells
_HYPERLINK_RE = re.compile(r'=HYPERLINK\("([^"]+)"', re.IGNORECASE)


# =============================================================================
# 1. Type coercion utilities
# =============================================================================

def _coerce_int(val: Any, fallback: int = 0) -> int:
    """Safely convert an Excel cell value to int."""
    try:
        if val is None or str(val).strip() in ("", "None"):
            return fallback
        return int(float(val))
    except (ValueError, TypeError):
        return fallback


def _coerce_bool(val: Any) -> bool:
    """Convert an Excel cell value to bool.  Handles 0/1, True/False, Yes/No."""
    if val is None:
        return False
    return str(val).strip().lower() in ("1", "true", "yes", "1.0")


def _int_val(raw: Any) -> int:
    """Alias used internally by parse_domain_excel."""
    return _coerce_int(raw)


def _float_val(raw: Any) -> float:
    try:
        return float(raw) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


# =============================================================================
# 2. Excel cell utilities
# =============================================================================

def _strip_hyperlink(val: Any) -> str:
    """Extract the bare URL from an Excel =HYPERLINK(...) formula, or return as-is."""
    if val is None:
        return ""
    s = str(val).strip()
    m = _HYPERLINK_RE.match(s)
    return m.group(1) if m else s


# =============================================================================
# 3. Filename / path utilities
# =============================================================================

def _parse_timestamp(filename: str) -> datetime | None:
    """Extract the datetime embedded in a report filename, or return None."""
    m = _TIMESTAMP_RE.search(filename)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), config.EXCEL_REPORT_TIMESTAMP_FORMAT)
    except ValueError:
        return None


def find_latest_xlsx(folder: Path) -> Path | None:
    """Return the most recently timestamped .xlsx in *folder*, or None."""
    candidates = [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")]
    return max(candidates, key=lambda p: _TIMESTAMP_RE.search(p.name).group(1)
               if _TIMESTAMP_RE.search(p.name) else "", default=None)


def folder_to_display_name(folder_name: str) -> str:
    """Convert an OneDrive folder name to a readable domain string.

    'calstatela-edu_ecst' → 'calstatela.edu_ecst'
    Only the portion before the first '_' has dashes converted to dots.
    """
    if "_" in folder_name:
        domain_part, rest = folder_name.split("_", 1)
        return f"{domain_part.replace('-', '.')}_{rest}"
    return folder_name.replace("-", ".")


def xlsx_report_date(xlsx_path: Path) -> str:
    """Return a human-readable scan date from an Excel filename.

    e.g. 'calstatela-edu_ecst-2026-03-20_08-14-32.xlsx' → 'March 20, 2026'
    Falls back to the filename stem if no timestamp is found.
    """
    ts = _parse_timestamp(xlsx_path.name)
    return ts.strftime("%B %d, %Y") if ts else xlsx_path.stem


# =============================================================================
# 4. Excel row reader  (raw rows — used by email draft builder)
# =============================================================================

def read_pdf_rows(xlsx_path: Path) -> list[dict]:
    """Read the 'Unique PDFs' sheet and return a list of row dicts keyed by header.

    Returns raw rows (one per Excel row, no deduplication).  Use
    parse_domain_excel() when you need aggregate metrics with deduplication.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        print(f"    [WARN] Cannot open {xlsx_path.name}: {exc}")
        return []

    sheet_name = (
        "Unique PDFs" if "Unique PDFs" in wb.sheetnames
        else "Scanned PDFs" if "Scanned PDFs" in wb.sheetnames
        else None
    )
    if sheet_name is None:
        wb.close()
        return []

    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result  = []
    for row in rows[1:]:
        if not any(cell is not None for cell in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


# =============================================================================
# 5. Row-level priority helper  (for per-PDF display in email drafts)
# =============================================================================

def row_to_priority_data(row: dict) -> dict:
    """Build the dict expected by filters.get_priority_level() from an Excel row.

    NOTE: this is used only for extracting display metrics (violations,
    page_count, etc.) in email drafts.  The HIGH/LOW priority classification
    in all reports and emails uses the pre-stamped 'Low Priority' column, not
    a live call to get_priority_level().
    """
    tagged = 1 if _coerce_bool(row.get("tagged")) else 0

    # failed_checks may have been serialised as Yes/No by a known export quirk.
    # Reconstruct from Errors/Page × page_count when that happens.
    failed_raw = row.get("failed_checks")
    if str(failed_raw).strip().lower() in ("yes", "no", ""):
        errors_per_page = _coerce_int(row.get("Errors/Page", 0))
        page_count      = _coerce_int(row.get("page_count", 0))
        failed_checks   = errors_per_page * page_count if page_count > 0 else 0
    else:
        failed_checks = _coerce_int(failed_raw)

    return {
        "tagged":               tagged,
        "pdf_text_type":        str(row.get("pdf_text_type") or ""),
        "violations":           _coerce_int(row.get("violations", 0)),
        "failed_checks":        failed_checks,
        "page_count":           _coerce_int(row.get("page_count", 0)),
        "has_form":             1 if _coerce_bool(row.get("has_form")) else 0,
        "approved_pdf_exporter": _coerce_bool(row.get("approved_pdf_exporter")),
    }


# =============================================================================
# 6. Aggregate Excel parser  ← THE canonical count source for all reports
# =============================================================================

def _col_index(header_row: tuple) -> dict[str, int]:
    return {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}


def parse_domain_excel(source: Path | bytes) -> dict | None:
    """Parse one domain Excel report and return aggregate metrics.

    *source* may be a file Path or raw bytes (Teams/Graph API downloads).
    Returns None when the file cannot be opened or has no usable data rows.

    All counts are fingerprint-deduplicated.  The 'Low Priority' column is the
    sole source of truth for high-priority classification (see module docstring).

    Returned dict keys
    ------------------
    unique_pdfs          int   — deduplicated PDF count
    high_priority        int   — PDFs where Low Priority == "No"
    compliant_scanned    int   — unique_pdfs − high_priority
    compliance_pct       float — compliant_scanned / unique_pdfs × 100
    violations_total     int   — sum of violations (deduplicated rows)
    violations_avg       float — mean violations per unique PDF
    errors_per_page_avg  float — mean Errors/Page (deduplicated rows)
    top_errors           dict  — {message: count} top-10 from Failure sheet
    total_scanned        int   — alias for unique_pdfs (backwards compat)
    """
    try:
        if isinstance(source, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        print(f"    [WARN] Cannot open workbook: {exc}")
        return None

    try:
        sheet_name = (
            "Unique PDFs"  if "Unique PDFs"  in wb.sheetnames else
            "Scanned PDFs" if "Scanned PDFs" in wb.sheetnames else None
        )
        if sheet_name is None:
            print(f"    [WARN] Neither 'Unique PDFs' nor 'Scanned PDFs' found "
                  f"(sheets: {wb.sheetnames})")
            return None

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if len(rows) < 2:
            print(f"    [WARN] '{sheet_name}' has no data rows")
            return None

        col    = _col_index(rows[0])
        v_idx  = col.get("violations")
        ep_idx = col.get("Errors/Page")
        lp_idx = col.get("Low Priority")
        fp_idx = col.get("fingerprint")

        # Deduplicate by fingerprint.  High-priority flag wins when the same
        # fingerprint appears more than once.
        seen: dict[str, dict] = {}
        for idx, row in enumerate(rows[1:]):
            if all(c is None for c in row):
                continue
            fp_raw = row[fp_idx] if fp_idx is not None else None
            key    = str(fp_raw).strip() if fp_raw is not None else f"__row_{idx}"
            if not key:
                key = f"__row_{idx}"

            raw_lp  = row[lp_idx] if lp_idx is not None else None
            is_high = isinstance(raw_lp, str) and raw_lp.strip().lower() == "no"
            v       = _int_val(row[v_idx]  if v_idx  is not None else None)
            ep      = _float_val(row[ep_idx] if ep_idx is not None else None)

            if key not in seen:
                seen[key] = {"is_high": is_high, "v": v, "ep": ep}
            else:
                if is_high:
                    seen[key]["is_high"] = True
                seen[key]["v"]  = v
                seen[key]["ep"] = ep

        unique_pdfs       = len(seen)
        high_priority     = sum(1 for r in seen.values() if r["is_high"])
        viol_list         = [r["v"]  for r in seen.values()]
        ep_list           = [r["ep"] for r in seen.values()]
        compliant_scanned = unique_pdfs - high_priority
        compliance_pct    = round(compliant_scanned / unique_pdfs * 100, 1) if unique_pdfs else 0.0
        violations_total  = sum(viol_list)
        violations_avg    = round(violations_total / unique_pdfs, 1) if unique_pdfs else 0.0
        errors_per_page_avg = round(sum(ep_list) / len(ep_list), 2) if ep_list else 0.0

        # Top errors from Failure sheet
        top_errors: dict[str, int] = {}
        if "Failure" in wb.sheetnames:
            f_rows = list(wb["Failure"].iter_rows(values_only=True))
            if len(f_rows) > 1:
                f_col   = _col_index(f_rows[0])
                msg_idx = f_col.get("error_message")
                if msg_idx is not None:
                    counter: Counter = Counter()
                    for row in f_rows[1:]:
                        if all(c is None for c in row):
                            continue
                        msg = row[msg_idx]
                        if msg:
                            counter[str(msg)[:150]] += 1
                    top_errors = dict(counter.most_common(10))

        return dict(
            unique_pdfs=unique_pdfs,
            high_priority=high_priority,
            compliant_scanned=compliant_scanned,
            compliance_pct=compliance_pct,
            violations_total=violations_total,
            violations_avg=violations_avg,
            errors_per_page_avg=errors_per_page_avg,
            total_scanned=unique_pdfs,   # backwards compat alias
            top_errors=top_errors,
        )

    except Exception as exc:
        print(f"    [WARN] Error parsing workbook: {exc}")
        return None
    finally:
        wb.close()


# =============================================================================
# 7. Domain scan collector
# =============================================================================

def collect_from_local(
    base_path: Path,
    domains: list[str] | None,
) -> dict[str, list[dict]]:
    """Read all timestamped Excel reports from each domain subfolder.

    Returns ``{display_key: [scan_dict, ...]}`` sorted oldest → newest.

    *domains* — None to auto-discover every subfolder that has xlsx files;
    or a list of domain keys (config naming) to process only those.
    """
    if domains is not None:
        folder_lookup = {
            config.get_domain_folder_name(d).lower(): config.get_domain_folder_name(d)
            for d in domains
        }
    else:
        folder_lookup = None

    data: dict[str, list[dict]] = {}

    for entry in sorted(base_path.iterdir()):
        if not entry.is_dir():
            continue

        if folder_lookup is not None:
            if entry.name.lower() not in folder_lookup:
                continue
            display_key = folder_lookup[entry.name.lower()]
        else:
            display_key = entry.name

        scans: list[dict] = []
        for xlsx in sorted(entry.glob("*.xlsx")):
            if xlsx.name.startswith("~$"):
                continue
            ts = _parse_timestamp(xlsx.name)
            if ts is None:
                print(f"    [SKIP] {xlsx.name} — no timestamp in filename")
                continue
            print(f"    [READ] {xlsx.name}")
            metrics = parse_domain_excel(xlsx)
            if metrics:
                metrics["timestamp"] = ts
                metrics["filename"]  = xlsx.name
                metrics["rel_path"]  = str(xlsx.relative_to(base_path))
                scans.append(metrics)
            else:
                print(f"    [SKIP] {xlsx.name} — could not extract metrics")

        if scans:
            scans.sort(key=lambda s: s["timestamp"])
            data[display_key] = scans

    return data


# =============================================================================
# 8. HTML / template utilities
# =============================================================================

def _js(obj: Any) -> str:
    """Serialize *obj* to compact JSON safe for embedding in a <script> block.

    Escapes <, >, & so that substrings like </script> or <!-- in error-message
    text cannot silently break the enclosing HTML <script> tag.
    """
    return (
        json.dumps(obj, default=str)
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )


# =============================================================================
# 9. Employee loader  (shared by send_emails.py and sharepoint_sync.py)
# =============================================================================

def load_employees(csv_path: Path) -> dict:
    """Return {employee_id: {first_name, last_name, email}}.

    employees.csv format (with header row):
        Full Name, Employee ID, Email
    """
    employees: dict = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header
        for row in reader:
            if not row or not row[1].strip():
                continue
            parts = row[0].strip().split(" ", 1)
            employees[row[1].strip()] = {
                "first_name": parts[0],
                "last_name": parts[1] if len(parts) > 1 else "",
                "email": row[2].strip() if len(row) > 2 else "",
            }
    return employees
