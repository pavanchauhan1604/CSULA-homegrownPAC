
import os
import sys
import re
import html
import shutil
import zipfile
from datetime import datetime

import requests
from urllib.parse import unquote
from openpyxl import load_workbook

# Add project root to path for imports
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from src.core.conformance_checker import loop_through_files_in_folder
from src.data_management.data_export import get_pdf_reports_by_site_name
from src.data_management.data_import import get_site_id_by_domain_name, mark_pdf_as_removed
from crawlers.sf_state_pdf_scan.sf_state_pdf_scan.box_handler import download_from_box, box_share_pattern_match


pdf_sites_folder = "C:\\Users\\913678186\\Box\\ATI\\PDF Accessibility\\SF State Website PDF Scans"
scans_output = "C:\\Users\\913678186\\Box\\ATI\\PDF Accessibility\\SF State Website PDF Scans\\{}"


def delete_scans_files(root_folder):
    # Walk through the directory tree
    for current_path, dirs, files in os.walk(root_folder):
        for file in files:
            # Check if the file name contains the substring
            if "_scans.xlsx" in file:
                file_path = os.path.join(current_path, file)
                try:
                    os.remove(file_path)
                    print(f"Deleted: {file_path}")
                except Exception as e:
                    print(f"Error deleting {file_path}: {e}")


def remove_timestamps_from_parent_urls():
    conn = sqlite3.connect('drupal_pdfs.db')
    cursor = conn.cursor()
    pdfs = cursor.execute("SELECT * FROM drupal_pdf_files").fetchall()

    for each in pdfs:
        original_url = each[2]
        print(each[0])
        # Remove everything after the first space (if any)
        cleaned_url = original_url.split(" ")[0]

        # If the URL has been changed, update the record
        if cleaned_url != original_url:
            print(f"Updating: {original_url} -> {cleaned_url}")
            # Assuming 'id' is the primary key and is stored in the first column.
            cursor.execute("UPDATE drupal_pdf_files SET parent_uri = ? WHERE id = ?", (cleaned_url, each[0]))

    conn.commit()
    conn.close()

def strip_trailing_items_from_pdf_urls():
    conn = sqlite3.connect('drupal_pdfs.db')
    cursor = conn.cursor()
    pdfs = cursor.execute("SELECT * FROM drupal_pdf_files").fetchall()

    for each in pdfs:
        original_url = each[1]
        # print(each[1])
        # Remove everything after the first space (if any)
        cleaned_url = original_url.split(" ")[0]

        # If the URL has been changed, update the record
        if cleaned_url != original_url:
            print(f"Updating: {original_url} -> {cleaned_url}")
            # Assuming 'id' is the primary key and is stored in the first column.
            cursor.execute("UPDATE drupal_pdf_files SET pdf_uri = ? WHERE id = ?", (cleaned_url, each[0]))

    conn.commit()
    conn.close()

import sqlite3

def delete_duplicate_entries():
    """
    For each site in the drupal_site table, this function reads the SQL query from
    delete_duplicates.sql (which deletes duplicate PDF entries for a given site by keeping only the oldest scan),
    substitutes the {site_name} placeholder with the actual domain name, and executes the query.

    The site names are retrieved by executing the SQL query stored in get_all_sites.sql.
    """
    # Connect to the SQLite database.
    conn = sqlite3.connect('drupal_pdfs.db')
    cursor = conn.cursor()

    # Read the get_all_sites.sql file, which contains:
    # select domain_name from drupal_site;
    with open("sql/get_all_sites.sql", "r") as f:
        get_sites_query = f.read().strip()

    # Execute the query to fetch all site domain names.
    cursor.execute(get_sites_query)
    sites = cursor.fetchall()  # Each row is a tuple (domain_name,)

    # Read the delete_duplicates.sql file which contains our delete query template.
    with open("sql/delete_duplicates.sql", "r") as f:
        delete_query_template = f.read()

    # Loop through each site and execute the delete query with the proper substitution.
    for site in sites:
        domain_name = site[0]
        formatted_query = delete_query_template.replace("{site_name}", domain_name)
        print(f"Deleting duplicate entries for site: {domain_name}")
        cursor.executescript(formatted_query)
        conn.commit()
        print(f"Finished processing site: {domain_name}")

    # Close the database connection.
    conn.close()


def download_all_dprc_will_remediate_pdfs_by_site(site_name):
    box_folder = rf'C:\Users\913678186\Box\ATI\PDF Accessibility\SF State Website PDF Scans\{site_name}'
    box_temp_folder = os.path.join(box_folder, 'temp')
    # Reports are now generated with timestamped names; pick the most recent.
    xlsx_candidates = [
        os.path.join(box_folder, f)
        for f in os.listdir(box_folder)
        if f.lower().endswith('.xlsx') and os.path.isfile(os.path.join(box_folder, f))
    ]
    xlsx_file = max(xlsx_candidates, key=os.path.getmtime) if xlsx_candidates else None

    if not os.path.exists(box_folder):
        print(f"File does not exist: {box_folder}")
        return
    print(f"Found folder: {box_folder}")

    if not xlsx_file:
        print(f"No .xlsx report found in: {box_folder}")
        return
    print(f"Using report: {xlsx_file}")

    # Compile hyperlink regex
    hyperlink_pattern = re.compile(r'HYPERLINK\("([^"]+)"')

    # Load workbook and sheet
    workbook = load_workbook(xlsx_file)
    sheet = workbook['Scanned PDFs']
    print(f"Processing sheet: {sheet.title}")

    # Prepare temp folder
    os.makedirs(box_temp_folder, exist_ok=True)

    # Download loop
    for row in sheet.iter_rows(min_row=1, max_col=17, max_row=sheet.max_row):
        link = row[0].value
        high_priority = row[15].value
        dprc_remediation = row[16].value

        if dprc_remediation == "Yes" and isinstance(link, str) and link.startswith('=HYPERLINK'):
            match = hyperlink_pattern.search(link)
            if not match:
                continue

            first_url = match.group(1)
            print(first_url, high_priority, dprc_remediation)

            if box_share_pattern_match(first_url):
                print("Downloading file from box…")
                download_from_box(first_url, box_temp_folder)
            else:
                print("Downloading file from URL…")
                response = requests.get(first_url, stream=True)
                if response.status_code == 200:
                    raw_name = os.path.basename(first_url)
                    unescaped = html.unescape(raw_name)
                    file_name = unquote(unescaped)
                    file_path = os.path.join(box_temp_folder, file_name)
                    with open(file_path, "wb") as file:
                        for chunk in response.iter_content(chunk_size=8192):
                            file.write(chunk)

    # --- ZIP and cleanup ---
    # Define the ZIP filename
    zip_name = f"{site_name}-pdf-scans.zip"
    zip_path = os.path.join(box_folder, zip_name)

    print(f"Creating ZIP archive: {zip_path}")
    with zipfile.ZipFile(zip_path, 'w', ZIP_DEFLATED:=zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(box_temp_folder):
            for fname in files:
                if fname.lower().endswith('.pdf'):
                    full_path = os.path.join(root, fname)
                    # Store PDFs at the root of the archive, no temp/ prefix
                    arcname = os.path.relpath(full_path, box_temp_folder)
                    zipf.write(full_path, arcname)

    print("ZIP archive created successfully.")

    # Remove the temp folder
    print(f"Removing temporary folder: {box_temp_folder}")
    shutil.rmtree(box_temp_folder)
    print("Temporary folder deleted.")


def get_all_folders_by_date_modified(folder_path, date_modified):
    """
    Returns a list of all folder names (not full paths) in the specified directory that were modified after the given date.

    Parameters:
        folder_path (str): The path to the directory to search.
        date_modified (str): The date (YYYY-MM-DD or MM/DD/YYYY) to compare against.

    Returns:
        list: A list of folder names that were modified after the specified date.
    """
    # Try parsing date in multiple formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            date_obj = datetime.strptime(date_modified, fmt)
            break
        except ValueError:
            continue
    else:
        raise ValueError("date_modified must be in YYYY-MM-DD or MM/DD/YYYY format")

    modified_folders = []
    for root, dirs, _ in os.walk(folder_path):
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            if os.path.getmtime(dir_path) > date_obj.timestamp():
                modified_folders.append(dir_name)
    return modified_folders





# print('\n'.join(get_all_folders_by_date_modified(
#     r"C:\Users\913678186\Box\ATI\PDF Accessibility\SF State Website PDF Scans",
#     "06/17/2025")))

# download_all_dprc_will_remediate_pdfs_by_site('cob-sfsu-edu')


def mark_pdfs_as_removed(site_folders):
    """
    Compare the raw pdf scrape URLS and Parent with current PDFS and mark current PDFS as removed if they are not in the raw scrape.
    :return:
    """
    raw_pdf_scan_set = set()

    for folder in os.listdir(site_folders):

        domain_id = get_site_id_by_domain_name(folder)

        site_pdfs = get_pdf_reports_by_site_name(folder.replace("-", "."))

        existing_pdfs_set = set((pdf.pdf_uri, pdf.parent_uri) for pdf in site_pdfs)

        if domain_id is not None:
            pdf_locations = loop_through_files_in_folder(os.path.join(site_folders, folder))

            if pdf_locations:

                for file in pdf_locations:

                    file_split = file.split(' ', 1)  # Splits at the last space

                    file_url = file_split[0]
                    loc = file_split[1].split(" ")[0]
                    raw_pdf_scan_set.add((file_url, loc))

        missing_pdfs = existing_pdfs_set.difference(raw_pdf_scan_set)
        if missing_pdfs:
            for pdf_uri, parent_uri in missing_pdfs:

                mark_pdf_as_removed(pdf_uri, parent_uri)


        raw_pdf_scan_set.clear()
        existing_pdfs_set.clear()


if __name__ == "__main__":
    mark_pdfs_as_removed(pdf_sites_folder)

