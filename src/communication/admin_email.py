#!/usr/bin/env python3
import sqlite3
import os
import re
from collections import defaultdict

import win32com.client
from openpyxl import load_workbook


def sanitize_filename(filename):
    # Remove characters that are invalid in file names.
    return re.sub(r'[\\/:"*?<>|]+', "", filename)

def run_sql_query(sql_file_path, template_html):
    # Open and read the SQL file.
    try:
        with open(sql_file_path, 'r') as file:
            sql_query = file.read()
    except Exception as e:
        print(f"Error reading SQL file: {e}")
        return

    # Connect to the SQLite database.
    try:
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
    except Exception as e:
        print(f"Error connecting to the database: {e}")
        return

    # Execute the SQL query and fetch results.
    try:
        cursor.execute(sql_query)
        results = cursor.fetchall()
    except Exception as e:
        print(f"Error executing SQL query: {e}")
        conn.close()
        return
    finally:
        conn.close()

    if not results:
        print("No rows returned by the query.")
        return

    # Prepare the output folder for .msg files.
    output_folder = r"C:\Users\913678186\Box\ATI\PDF Accessibility\MPP Emails"
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Create an instance of Outlook.
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
    except Exception as e:
        print(f"Error initializing Outlook: {e}")
        return

    # Process each row from the query results.
    for row in results:
        # Ensure row has at least 4 columns and skip if index 1 is None or an empty string.
        if len(row) < 4 or row[1] is None or row[1] == "":
            continue

        first_name = row[1]
        recipient_email = row[3]

        # Format the HTML content by replacing the {site_manager} placeholder.
        html_content = template_html.format(site_manager=first_name)
        print(html_content)
        try:
            # Create a new MailItem.
            mail_item = outlook.CreateItem(0)  # 0 indicates a MailItem.
            mail_item.To = recipient_email
            mail_item.SentOnBehalfOfName = "access@sfsu.edu"
            mail_item.Subject = "SF State ATI | Drupal PDF Accessibility Remediation"
            mail_item.HTMLBody = html_content

            # Create a safe filename based on first name and recipient email.
            filename = sanitize_filename(f"{first_name}_{recipient_email}.msg")
            file_path = os.path.join(output_folder, filename)

            # Save the mail item as a .msg file (3 represents the olMSG format).
            mail_item.SaveAs(file_path, 3)
            print(f"Saved MSG file: {file_path}")
        except Exception as e:
            print(f"Error processing row {row}: {e}")




def generate_followup_emails(excel_sheet, sql_file_path, template_html):
    # Load the Excel workbook
    # Load the Excel workbook
    try:
        wb = load_workbook(excel_sheet, read_only=True)
        if "Third Round Followup" not in wb.sheetnames:
            print("Sheet 'Third Round Followup' not found in the Excel file.")
            return
        sheet = wb["Third Round Followup"]
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Group domains by email address (no headers, so use fixed column indices)
    email_to_domains = defaultdict(set)

    try:
        for row in sheet.iter_rows(min_row=1, values_only=True):  # Start at first row
            domain = row[0]  # Column A
            email = row[2]   # Column C
            if email and domain and isinstance(email, str):
                print(email, domain)
                email_to_domains[email.strip()].add(domain.strip())
    except Exception as e:
        print(f"Error processing Excel data: {e}")
        return

    # Load SQL query
    try:
        with open(sql_file_path, 'r') as f:
            sql_query = f.read()
    except Exception as e:
        print(f"Error reading SQL file: {e}")
        return

    try:
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
        cursor.execute(sql_query)
        results = cursor.fetchall()
        print("results", results)
    except Exception as e:
        print(f"Error executing SQL query: {e}")
        return
    finally:
        conn.close()

    if not results:
        print("No results returned from SQL.")
        return

    # Map emails to names from SQL (email = row[2], name = row[1])
    email_to_name = {}
    print(email_to_domains)
    for row in results:
        print("ROW", row)
        if len(row) < 4 or not row[2]:
            continue
        email = row[3].strip()
        name = row[1]
        print("EMAIL", email)
        if email in email_to_domains:
            email_to_name[email] = name
    print("DD", email_to_name)
    # Initialize Outlook
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
    except Exception as e:
        print(f"Error initializing Outlook: {e}")
        return

    # Output folder
    output_folder = r"C:\Users\913678186\Box\ATI\PDF Accessibility\Follow-Up Emails"
    os.makedirs(output_folder, exist_ok=True)

    # Generate one email per unique address
    for email, domains in email_to_domains.items():
        if email not in email_to_name:
            continue  # Skip if email not matched in SQL query

        name = email_to_name[email]
        domain_list = sorted(domains)
        domain_str = "<ul>" + "".join(f"<li>{domain}</li>" for domain in domain_list) + "</ul>"

        try:
            html_content = template_html.format(site_manager=name, domain_list=domain_str)

            mail_item = outlook.CreateItem(0)
            mail_item.To = email
            mail_item.SentOnBehalfOfName = "access@sfsu.edu"
            mail_item.Subject = "SF State ATI | Follow-Up: Drupal PDF Accessibility"
            mail_item.HTMLBody = html_content

            filename = sanitize_filename(f"{name}_{email}_followup.msg")
            file_path = os.path.join(output_folder, filename)
            mail_item.SaveAs(file_path, 3)

            print(f"Saved follow-up MSG: {file_path}")
        except Exception as e:
            print(f"Error processing email to {email}: {e}")



if __name__ == "__main__":
    # Path to the SQL file with your query.
    sql_file_path = r"C:\Users\913678186\IdeaProjects\sf_state_pdf_website_scan\sql\get_admin_contacts.sql"

    # Path to the HTML template file.
    template_html_path = r"C:\Users\913678186\Box\ATI\PDF Accessibility\Reports\build_files\templates\mpp_followup_contact_email.html"

    # Open and read the HTML template file.
    try:
        with open(template_html_path, 'r', encoding='utf-8') as f:
            template_html = f.read()
    except Exception as e:
        print(f"Error reading HTML template file: {e}")
        exit(1)

    generate_followup_emails(r"C:\Users\913678186\Box\ATI\PDF Accessibility\PDF Project Documents\drupal_pdf_working_group.xlsx",
                             # sql_file_path,
                             # template_html)
