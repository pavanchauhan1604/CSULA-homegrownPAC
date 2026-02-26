"""Historical PDF accessibility trend analysis.

Fetches every timestamped .xlsx report for each domain from the Teams channel
(or a local folder), parses key metrics, and produces a self-contained HTML
dashboard with Chart.js charts — similar in spirit to the PopeTech trend view
but scoped to unique-PDF counts, compliance rates, error breakdowns, and
high-priority counts across all your historical scans.

Metrics extracted from each report
-----------------------------------
Scanned PDFs sheet:
  - Total rows (all PDF occurrences across all pages)
  - Compliance % (Low Priority PDFs / total unique PDFs — PDFs where Low Priority == "Yes")
  - Average Errors/Page
  - High Priority count (Low Priority == "No")

Unique PDFs sheet:
  - Unique PDF count (deduplicated by file)

Failure sheet:
  - Top 10 error message types by occurrence count

Usage
-----
    # Generate a per-domain HTML report for every domain (saves into each domain folder):
    python scripts/historical_analysis.py

    # Specific domains only:
    python scripts/historical_analysis.py --domains www.calstatela.edu_admissions libguides.calstatela.edu

    # Read from a local folder instead of OneDrive:
    python scripts/historical_analysis.py --source local --local-path "C:/Downloads/TeamsFiles"

    # Generate locally only (don't write back to OneDrive domain folders):
    python scripts/historical_analysis.py --no-upload
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import config
import openpyxl

# Regex that matches the timestamp embedded in report filenames
# e.g.  www.calstatela.edu_admissions-2026-01-25_06-26-57.xlsx
_TIMESTAMP_RE = re.compile(
    r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE
)

# Chart.js colour palette — CSULA navy first, then accessible contrasting colours
_CHART_COLORS = [
    "#003262", "#C4820E", "#27ae60", "#e67e22",
    "#8e44ad", "#16a085", "#c0392b", "#2980b9",
    "#229954", "#d4ac0d",
]


# =============================================================================
# Excel parsing
# =============================================================================

def _col_index(header_row: tuple) -> dict[str, int]:
    return {
        str(h).strip(): i
        for i, h in enumerate(header_row)
        if h is not None
    }


def _int_val(raw: Any) -> int:
    try:
        return int(raw) if raw is not None else 0
    except (ValueError, TypeError):
        return 0


def _float_val(raw: Any) -> float:
    try:
        return float(raw) if raw is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def parse_excel_report(source: Path | bytes) -> dict | None:
    """Parse one Excel accessibility report and return a metrics dict.

    *source* can be a file Path or raw bytes (when downloaded from Teams).
    Returns None when the file cannot be read or contains no data rows.
    """
    try:
        if isinstance(source, (bytes, bytearray)):
            wb = openpyxl.load_workbook(
                io.BytesIO(source), read_only=True, data_only=False
            )
        else:
            wb = openpyxl.load_workbook(source, read_only=True, data_only=False)
    except Exception as exc:
        print(f"    [WARN] Cannot open workbook: {exc}")
        return None

    result: dict[str, Any] = {}

    # ── Unique PDFs sheet — all metrics are based on deduplicated PDFs ────
    # Fall back to Scanned PDFs only if Unique PDFs sheet is absent (old reports).
    sheet_name = "Unique PDFs" if "Unique PDFs" in wb.sheetnames else "Scanned PDFs"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None

    rows = list(wb[sheet_name].iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return None

    col = _col_index(rows[0])
    v_idx  = col.get("violations")
    ep_idx = col.get("Errors/Page")
    lp_idx = col.get("Low Priority")

    unique_pdfs = 0
    compliant = 0
    violations_list: list[int] = []
    ep_list: list[float] = []
    high_priority = 0

    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        unique_pdfs += 1

        v = _int_val(row[v_idx] if v_idx is not None else None)
        violations_list.append(v)
        if v == 0:
            compliant += 1

        if ep_idx is not None:
            ep_list.append(_float_val(row[ep_idx]))

        if lp_idx is not None:
            raw_lp = row[lp_idx]
            if isinstance(raw_lp, str) and raw_lp.strip().lower() == "no":
                high_priority += 1

    low_priority_pdfs = unique_pdfs - high_priority
    result.update(
        total_scanned=unique_pdfs,
        unique_pdfs=unique_pdfs,
        compliant_scanned=low_priority_pdfs,
        compliance_pct=round(low_priority_pdfs / unique_pdfs * 100, 1) if unique_pdfs else 0.0,
        violations_total=sum(violations_list),
        violations_avg=round(sum(violations_list) / len(violations_list), 1)
        if violations_list
        else 0.0,
        high_priority=high_priority,
        errors_per_page_avg=round(sum(ep_list) / len(ep_list), 2) if ep_list else 0.0,
    )

    # ── Failure sheet ─────────────────────────────────────────────────────
    top_errors: dict[str, int] = {}
    if "Failure" in wb.sheetnames:
        f_rows = list(wb["Failure"].iter_rows(values_only=True))
        if len(f_rows) > 1:
            f_col = _col_index(f_rows[0])
            msg_idx = f_col.get("error_message")
            if msg_idx is not None:
                counter: Counter = Counter()
                for row in f_rows[1:]:
                    if all(c is None for c in row):
                        continue
                    msg = row[msg_idx]
                    if msg:
                        counter[str(msg)[:150]] += 1
                top_errors = dict(counter.most_common(10))

    result["top_errors"] = top_errors
    wb.close()
    return result


# =============================================================================
# Data collection helpers
# =============================================================================

def _parse_timestamp(filename: str) -> datetime | None:
    m = _TIMESTAMP_RE.search(filename)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), config.EXCEL_REPORT_TIMESTAMP_FORMAT)
    except ValueError:
        return None


def collect_from_local(base_path: Path, domains: list[str] | None) -> dict[str, list[dict]]:
    """Read scan reports from a local folder that mirrors the Teams structure.

    When *domains* is None, every subdirectory that contains timestamped .xlsx
    files is processed automatically (auto-discover mode).  When a list of
    domain keys is given, only matching folders are processed (case-insensitive).
    """
    # Build a lower-case lookup: folder_name_lower → preferred display key
    if domains is not None:
        folder_to_domain = {
            config.get_domain_folder_name(d).lower(): config.get_domain_folder_name(d)
            for d in domains
        }
    else:
        folder_to_domain = None  # auto-discover all folders

    data: dict[str, list[dict]] = {}

    for entry in sorted(base_path.iterdir()):
        if not entry.is_dir():
            continue

        if folder_to_domain is not None:
            # Filter mode: skip folders not in the requested domain list
            if entry.name.lower() not in folder_to_domain:
                continue
            display_key = folder_to_domain[entry.name.lower()]
        else:
            # Auto-discover mode: use the folder name as-is
            display_key = entry.name

        scans: list[dict] = []
        for xlsx in sorted(entry.glob("*.xlsx")):
            if xlsx.name.startswith("~$"):
                continue
            ts = _parse_timestamp(xlsx.name)
            if ts is None:
                continue
            metrics = parse_excel_report(xlsx)
            if metrics:
                metrics["timestamp"] = ts
                metrics["filename"] = xlsx.name
                metrics["rel_path"] = str(xlsx.relative_to(base_path))
                scans.append(metrics)

        if scans:
            scans.sort(key=lambda s: s["timestamp"])
            data[display_key] = scans

    return data


def collect_from_teams(domains: list[str]) -> dict[str, list[dict]]:
    """Fetch scan reports from the Teams channel via Graph API."""
    from src.utilities.ms_graph import (
        get_token, get_drive_id,
        list_folder_children, list_item_children,
        download_item,
    )

    _check_teams_config()

    print("Authenticating with Microsoft Graph…")
    token = get_token(
        config.TEAMS_CLIENT_ID,
        config.TEAMS_TENANT_ID,
        config.TEAMS_TOKEN_CACHE_PATH,
    )
    print("Authenticated.\n")

    print("Resolving SharePoint drive…")
    drive_id = get_drive_id(token, config.TEAMS_SHAREPOINT_SITE_URL)
    print(f"Drive ID: {drive_id}\n")

    folder_to_domain = {config.get_domain_folder_name(d): d for d in domains}
    data: dict[str, list[dict]] = {}

    print(f"Listing subfolders under '{config.TEAMS_FOLDER_PATH}'…")
    try:
        root_items = list_folder_children(token, drive_id, config.TEAMS_FOLDER_PATH)
    except Exception as exc:
        print(f"ERROR: Could not list Teams folder '{config.TEAMS_FOLDER_PATH}': {exc}")
        return {}

    for item in root_items:
        if "folder" not in item:
            continue
        folder_name = item["name"]
        domain = folder_to_domain.get(folder_name)
        if domain is None:
            continue

        print(f"\n  ▶ {domain}")
        children = list_item_children(token, drive_id, item["id"])

        scans: list[dict] = []
        for child in children:
            if "file" not in child:
                continue
            fname = child["name"]
            if not fname.lower().endswith(".xlsx") or fname.startswith("~$"):
                continue
            ts = _parse_timestamp(fname)
            if ts is None:
                continue

            print(f"    Downloading {fname}…")
            try:
                file_bytes = download_item(token, child)
                metrics = parse_excel_report(file_bytes)
                if metrics:
                    metrics["timestamp"] = ts
                    metrics["filename"] = fname
                    scans.append(metrics)
            except Exception as exc:
                print(f"    [WARN] {fname}: {exc}")

        if scans:
            scans.sort(key=lambda s: s["timestamp"])
            data[domain] = scans
            print(f"    → {len(scans)} scan(s) loaded")
        else:
            print(f"    → no valid timestamped scans found")

    return data


# =============================================================================
# HTML report generation
# =============================================================================

def _js(obj: Any) -> str:
    """Serialize *obj* to compact JSON safe for embedding in a JS context."""
    return json.dumps(obj, default=str)


def _anchor(domain: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]", "-", domain)


def _sharepoint_url(rel_path: str) -> str:
    """Build a direct SharePoint open-in-browser URL for a relative file path."""
    base = getattr(config, "TEAMS_SHAREPOINT_FILES_URL", "").rstrip("/")
    if not base or not rel_path:
        return ""
    encoded = "/".join(quote(part, safe="") for part in Path(rel_path).parts)
    return f"{base}/{encoded}?csf=1&web=1"


def _trend_arrow(scans: list[dict]) -> tuple[str, str]:
    """Return (display_text, css_class) for compliance trend."""
    if len(scans) < 2:
        return "— (1 scan)", "trend-flat"
    diff = scans[-1]["compliance_pct"] - scans[0]["compliance_pct"]
    if diff > 2:
        return f"↑ +{diff:.1f}%", "trend-up"
    elif diff < -2:
        return f"↓ {diff:.1f}%", "trend-down"
    return f"→ {diff:+.1f}%", "trend-flat"


def _build_error_datasets(scans: list[dict]) -> tuple[list[str], list[dict]]:
    """Build Chart.js datasets for top-error trend chart across all scans."""
    combined: Counter = Counter()
    for s in scans:
        combined.update(s["top_errors"])
    top_keys = [k for k, _ in combined.most_common(10)]

    date_labels = [s["timestamp"].strftime("%Y-%m-%d") for s in scans]
    datasets = []
    for i, key in enumerate(top_keys):
        short = key if len(key) <= 80 else key[:77] + "…"
        color = _CHART_COLORS[i % len(_CHART_COLORS)]
        datasets.append({
            "label": short,
            "data": [s["top_errors"].get(key, 0) for s in scans],
            "borderColor": color,
            "backgroundColor": color + "33",
            "tension": 0.3,
            "fill": False,
        })
    return date_labels, datasets


def _summary_row(domain: str, scans: list[dict]) -> str:
    latest = scans[-1]
    trend_text, trend_cls = _trend_arrow(scans)
    return (
        f'<tr>'
        f'<td><a href="#{_anchor(domain)}">{domain}</a></td>'
        f'<td class="c">{len(scans)}</td>'
        f'<td class="c">{scans[0]["timestamp"].strftime("%Y-%m-%d")}</td>'
        f'<td class="c">{latest["timestamp"].strftime("%Y-%m-%d")}</td>'
        f'<td class="c">{latest["unique_pdfs"]}</td>'
        f'<td class="c">{latest["compliance_pct"]:.1f}% ({latest["compliant_scanned"]}/{latest["unique_pdfs"]})</td>'
        f'<td class="c {trend_cls}">{trend_text}</td>'
        f'</tr>\n'
    )


def _domain_section(domain: str, scans: list[dict], idx: int) -> tuple[str, int]:
    """Return (html_string, next_chart_index) for one domain's section."""
    latest = scans[-1]
    multi = len(scans) >= 2

    date_labels = [s["timestamp"].strftime("%Y-%m-%d %H:%M") for s in scans]
    unique_data = [s["unique_pdfs"] for s in scans]
    comp_data   = [s["compliance_pct"] for s in scans]
    ep_data     = [s["errors_per_page_avg"] for s in scans]

    err_date_labels, err_datasets = _build_error_datasets(scans)

    pdfs_chart_id  = f"c-pdfs-{idx}"
    comp_chart_id  = f"c-comp-{idx}"
    ep_chart_id    = f"c-ep-{idx}"
    err_chart_id   = f"c-err-{idx}"
    idx += 1

    # Compliance card colour class
    pct = latest["compliance_pct"]
    card_pct_cls = "card-good" if pct >= 80 else ("card-warn" if pct >= 50 else "card-bad")

    # Scan history table rows (newest first)
    trows = ""
    for s in reversed(scans):
        ts_str = s["timestamp"].strftime("%Y-%m-%d %H:%M")
        sp_url = _sharepoint_url(s.get("rel_path", ""))
        ts_cell = (
            f'<a href="{sp_url}" target="_blank" class="xlsx-link">{ts_str}</a>'
            if sp_url else ts_str
        )
        trows += (
            f'<tr>'
            f'<td>{ts_cell}</td>'
            f'<td class="c">{s["unique_pdfs"]}</td>'
            f'<td class="c">{s["violations_total"]}</td>'
            f'<td class="c">{s["compliance_pct"]:.1f}% ({s["compliant_scanned"]}/{s["unique_pdfs"]})</td>'
            f'<td class="c">{s["errors_per_page_avg"]:.2f}</td>'
            f'<td class="c">{s["high_priority"]}</td>'
            f'</tr>\n'
        )

    if multi:
        # JavaScript chart initialisation block
        err_chart_block = ""
        if err_datasets:
            err_chart_block = (
                f'new Chart(document.getElementById("{err_chart_id}"), {{'
                f'  type:"line",'
                f'  data:{{ labels:{_js(err_date_labels)}, datasets:{_js(err_datasets)} }},'
                f'  options:{{ responsive:true, plugins:{{ legend:{{ position:"bottom", labels:{{ boxWidth:12, font:{{ size:11 }} }} }} }},'
                f'    scales:{{ y:{{ beginAtZero:true, title:{{ display:true, text:"Occurrences" }} }} }} }}'
                f'}});'
            )

        charts_block = f"""
        <div class="charts-grid">
          <div class="chart-box">
            <h4>PDF Inventory Over Time</h4>
            <canvas id="{pdfs_chart_id}"></canvas>
          </div>
          <div class="chart-box">
            <h4>Compliant (Low Priority) PDFs % Over Time</h4>
            <canvas id="{comp_chart_id}"></canvas>
          </div>
          <div class="chart-box">
            <h4>Average Errors per Page Over Time</h4>
            <canvas id="{ep_chart_id}"></canvas>
          </div>
          {"" if not err_datasets else f'<div class="chart-box chart-wide"><h4>Top Error Types Over Time</h4><canvas id="{err_chart_id}"></canvas></div>'}
        </div>
        <script>
        (function(){{
          new Chart(document.getElementById("{pdfs_chart_id}"),{{
            type:"line",
            data:{{ labels:{_js(date_labels)},
              datasets:[
                {{ label:"Unique PDFs", data:{_js(unique_data)}, borderColor:"#003262", backgroundColor:"#00326220", tension:0.3, fill:false }}
              ]
            }},
            options:{{ responsive:true, plugins:{{ legend:{{ position:"top" }} }}, scales:{{ y:{{ beginAtZero:true }} }} }}
          }});

          new Chart(document.getElementById("{comp_chart_id}"),{{
            type:"line",
            data:{{ labels:{_js(date_labels)},
              datasets:[
                {{ label:"Compliant PDFs %", data:{_js(comp_data)}, borderColor:"#27ae60", backgroundColor:"#27ae6018", tension:0.3, fill:true }},
                {{ label:"Target 100%",  data:{_js([100]*len(scans))}, borderColor:"#b0b8c4", borderDash:[6,4], pointRadius:0, fill:false }}
              ]
            }},
            options:{{ responsive:true, plugins:{{ legend:{{ position:"top" }} }}, scales:{{ y:{{ min:0, max:100 }} }} }}
          }});

          new Chart(document.getElementById("{ep_chart_id}"),{{
            type:"line",
            data:{{ labels:{_js(date_labels)},
              datasets:[
                {{ label:"Avg Errors/Page", data:{_js(ep_data)}, borderColor:"#C4820E", backgroundColor:"#C4820E18", tension:0.3, fill:true }}
              ]
            }},
            options:{{ responsive:true, plugins:{{ legend:{{ position:"top" }} }}, scales:{{ y:{{ beginAtZero:true }} }} }}
          }});

          {err_chart_block}
        }})();
        </script>"""
    else:
        charts_block = (
            '<p class="note">Only one scan on record — '
            "trend charts will appear once additional scans are uploaded.</p>"
        )

    html = f"""
<section id="{_anchor(domain)}" class="domain-section">
  <h2>{domain}</h2>
  <div class="cards">
    <div class="card"><span class="val">{latest["unique_pdfs"]}</span><span class="lbl">Unique PDFs<br>(latest)</span></div>
    <div class="card {card_pct_cls}"><span class="val">{latest["compliance_pct"]:.1f}%</span><span class="lbl">Compliant PDFs<br>({latest["compliant_scanned"]}/{latest["unique_pdfs"]})</span></div>
    <div class="card"><span class="val">{latest["violations_total"]}</span><span class="lbl">Total<br>Violations</span></div>
    <div class="card"><span class="val">{latest["errors_per_page_avg"]:.2f}</span><span class="lbl">Avg Errors<br>/Page</span></div>
    <div class="card card-warn"><span class="val">{latest["high_priority"]}</span><span class="lbl">High Priority<br>PDFs</span></div>
    <div class="card"><span class="val">{len(scans)}</span><span class="lbl">Scans<br>on Record</span></div>
  </div>
  {charts_block}
  <h4>Scan History</h4>
  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>Scan Date / Time</th><th>Unique PDFs</th>
        <th>Total Violations</th><th>Compliant PDFs (Low Priority)</th>
        <th>Avg Err/Page</th><th>High Priority</th>
      </tr></thead>
      <tbody>{trows}</tbody>
    </table>
  </div>
  <a href="#top" class="back">↑ Back to top</a>
</section>"""

    return html, idx


def build_html(domain_data: dict[str, list[dict]], generated_at: datetime) -> str:
    """Assemble the complete HTML dashboard string."""
    nav = "\n".join(
        f'<a href="#{_anchor(d)}">{d}</a>'
        for d in sorted(domain_data)
    )

    summary_rows = "".join(
        _summary_row(d, domain_data[d])
        for d in sorted(domain_data)
    )

    sections_html = ""
    chart_idx = 0
    for domain in sorted(domain_data):
        sec, chart_idx = _domain_section(domain, domain_data[domain], chart_idx)
        sections_html += sec

    total_domains = len(domain_data)
    total_scans   = sum(len(v) for v in domain_data.values())
    gen_str       = generated_at.strftime("%B %d, %Y at %I:%M %p")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CSULA PDF Accessibility — Historical Analysis</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,"Segoe UI",sans-serif;background:#f2f2f2;color:#1a1a1a;font-size:14px;line-height:1.5}}
header{{background:#111111;color:#fff;padding:26px 44px;border-bottom:4px solid #C4820E}}
header h1{{font-size:1.45rem;font-weight:700;letter-spacing:.01em}}
header p{{font-size:.82rem;color:#aaaaaa;margin-top:6px}}
nav{{background:#1a1a1a;border-bottom:2px solid #C4820E;padding:10px 44px;display:flex;flex-wrap:wrap;gap:16px;position:sticky;top:0;z-index:10}}
nav a{{color:#C4820E;text-decoration:none;font-size:.8rem;font-weight:600;white-space:nowrap;letter-spacing:.02em}}
nav a:hover{{color:#fff}}
.wrap{{max-width:1340px;margin:0 auto;padding:30px 44px}}
.pills{{display:flex;gap:14px;margin-bottom:30px;flex-wrap:wrap}}
.pill{{background:#fff;border:1px solid #d0d0d0;border-top:3px solid #C4820E;border-radius:4px;padding:16px 26px;text-align:center;min-width:130px;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.pill .pv{{font-size:2rem;font-weight:700;color:#111111;line-height:1}}
.pill .pl{{font-size:.72rem;color:#666;margin-top:5px;text-transform:uppercase;letter-spacing:.05em}}
h2.sh{{font-size:.85rem;font-weight:700;margin-bottom:14px;color:#111111;text-transform:uppercase;letter-spacing:.08em;border-left:3px solid #C4820E;padding-left:10px}}
.tbl-wrap{{background:#fff;border-radius:4px;border:1px solid #d0d0d0;overflow:auto;margin-bottom:32px;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
table{{width:100%;border-collapse:collapse;font-size:.855rem}}
th{{background:#111111;color:#C4820E;padding:10px 14px;text-align:left;white-space:nowrap;font-weight:700;font-size:.78rem;text-transform:uppercase;letter-spacing:.06em}}
td{{padding:9px 14px;border-bottom:1px solid #ebebeb;color:#222}}
tr:last-child td{{border-bottom:none}}
tr:nth-child(even) td{{background:#fafafa}}
tr:hover td{{background:#fff8ee}}
.c{{text-align:center}}
.trend-up{{color:#1a6e3a;font-weight:700}}
.trend-down{{color:#8b0000;font-weight:700}}
.trend-flat{{color:#888}}
.domain-section{{background:#fff;border-radius:4px;border:1px solid #d0d0d0;border-top:4px solid #C4820E;padding:28px 32px;margin-bottom:28px;box-shadow:0 2px 6px rgba(0,0,0,.06)}}
.domain-section h2{{font-size:1.1rem;color:#111111;border-bottom:2px solid #C4820E;padding-bottom:10px;margin-bottom:20px;font-weight:700}}
.domain-section h4{{font-size:.75rem;font-weight:700;margin:20px 0 10px;color:#111111;text-transform:uppercase;letter-spacing:.07em;border-left:2px solid #C4820E;padding-left:8px}}
.cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px}}
.card{{border-radius:4px;border:1px solid #d0d0d0;border-bottom:3px solid #aaa;background:#fff;padding:15px 20px;min-width:118px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.05)}}
.card-good{{border-bottom-color:#1a6e3a}}
.card-warn{{border-bottom-color:#C4820E}}
.card-bad{{border-bottom-color:#8b0000}}
.val{{display:block;font-size:1.55rem;font-weight:700;color:#111111;line-height:1.15}}
.card-good .val{{color:#1a6e3a}}
.card-warn .val{{color:#7a4f00}}
.card-bad .val{{color:#8b0000}}
.lbl{{display:block;font-size:.7rem;color:#666;margin-top:5px;line-height:1.35}}
.charts-grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px}}
.chart-box{{border:1px solid #d0d0d0;border-radius:4px;padding:16px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.04)}}
.chart-box h4{{font-size:.75rem;color:#111111;margin-bottom:12px;text-transform:uppercase;letter-spacing:.06em;font-weight:700;border-left:2px solid #C4820E;padding-left:8px}}
.chart-wide{{grid-column:1/-1}}
.note{{color:#666;font-size:.84rem;font-style:italic;margin-bottom:18px;padding:12px 16px;background:#fafafa;border-radius:4px;border-left:3px solid #C4820E}}
.xlsx-link{{color:#7a4f00;text-decoration:none;font-weight:600;border-bottom:1px dotted #C4820E}}
.xlsx-link:hover{{color:#C4820E;border-bottom-style:solid}}
.back{{display:inline-block;margin-top:18px;font-size:.78rem;color:#C4820E;text-decoration:none;font-weight:600}}
.back:hover{{color:#111111}}
footer{{text-align:center;padding:24px;font-size:.75rem;color:#aaa;border-top:2px solid #C4820E;margin-top:10px;background:#111111}}
@media(max-width:768px){{.wrap{{padding:16px 18px}}.charts-grid{{grid-template-columns:1fr}}header,nav{{padding-left:18px;padding-right:18px}}}}
</style>
</head>
<body id="top">
<header>
  <h1>CSULA PDF Accessibility — Historical Analysis</h1>
  <p>Generated: {gen_str} &nbsp;·&nbsp; {total_domains} domain(s) &nbsp;·&nbsp; {total_scans} total scan(s)</p>
</header>
<nav>{nav}</nav>
<div class="wrap">
  <div class="pills">
    <div class="pill"><div class="pv">{total_domains}</div><div class="pl">Domains</div></div>
    <div class="pill"><div class="pv">{total_scans}</div><div class="pl">Total Scans</div></div>
  </div>
  <h2 class="sh">Domain Summary</h2>
  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>Domain</th><th>Scans</th><th>First Scan</th><th>Latest Scan</th>
        <th>Unique PDFs</th><th>Compliant PDFs (Low Priority)</th><th>Trend</th>
      </tr></thead>
      <tbody>{summary_rows}</tbody>
    </table>
  </div>
  {sections_html}
</div>
<footer>CSULA PDF Accessibility Checker &nbsp;·&nbsp; {generated_at.year}</footer>
</body>
</html>"""


# =============================================================================
# Config check helper
# =============================================================================

def _check_onedrive_config() -> Path:
    """Return the validated OneDrive root path from config."""
    raw = getattr(config, "TEAMS_ONEDRIVE_PATH", "")
    if not raw:
        print("ERROR: TEAMS_ONEDRIVE_PATH is not set in config.py")
        sys.exit(1)
    p = Path(raw)
    if not p.exists():
        print(f"ERROR: OneDrive path does not exist: {p}")
        print("Make sure your Teams channel Files folder is synced via OneDrive.")
        sys.exit(1)
    return p


# =============================================================================
# Entry point
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a historical PDF accessibility trend report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--domains", nargs="+", metavar="DOMAIN",
        help="Domain(s) to analyse (default: all in config.DOMAINS)",
    )
    parser.add_argument(
        "--source", choices=["onedrive", "local"], default="onedrive",
        help=(
            "Where to read reports from: "
            "'onedrive' (default — uses TEAMS_ONEDRIVE_PATH from config) or "
            "'local' (any folder via --local-path)."
        ),
    )
    parser.add_argument(
        "--local-path", metavar="PATH",
        help="Root folder containing domain subfolders (required when --source=local)",
    )
    parser.add_argument(
        "--no-upload", action="store_true", dest="no_upload",
        help=(
            "Save HTML reports to output/reports/<domain>/ locally instead of "
            "back into the OneDrive domain folders (only relevant when --source=onedrive)."
        ),
    )
    args = parser.parse_args()

    # None = auto-discover all subfolders; explicit list = filter to those domains only
    domains = args.domains or None

    if args.source == "onedrive":
        onedrive_path = _check_onedrive_config()
        print(f"Reading reports from OneDrive: {onedrive_path}\n")
        domain_data = collect_from_local(onedrive_path, domains)
    else:  # local
        if not args.local_path:
            parser.error("--local-path is required when --source=local")
        local_path = Path(args.local_path)
        if not local_path.exists():
            print(f"ERROR: Path does not exist: {local_path}")
            sys.exit(1)
        print(f"Reading reports from local path: {local_path}\n")
        domain_data = collect_from_local(local_path, domains)

    if not domain_data:
        print(
            "\nNo scan data found. "
            "Make sure the folder contains timestamped .xlsx reports "
            "matching the naming convention: {domain}-YYYY-MM-DD_HH-MM-SS.xlsx"
        )
        sys.exit(0)

    generated_at = datetime.now()
    report_filename = "historical_analysis.html"

    print(f"Generating reports for {len(domain_data)} domain(s)…\n")
    saved = 0

    for domain, scans in sorted(domain_data.items()):
        print(f"▶ {domain}  ({len(scans)} scan(s))")
        html = build_html({domain: scans}, generated_at)

        # In auto-discover mode, `domain` is already the folder name on disk.
        # In filtered mode it went through get_domain_folder_name(), so convert only when needed.
        folder_name = domain if args.domains is None else config.get_domain_folder_name(domain)

        if args.source == "onedrive" and not args.no_upload:
            out_path = onedrive_path / folder_name / report_filename
        else:
            local_out = config.OUTPUT_REPORTS_DIR / folder_name
            local_out.mkdir(parents=True, exist_ok=True)
            out_path = local_out / report_filename

        out_path.write_text(html, encoding="utf-8")
        print(f"  ✓ {out_path}")
        saved += 1

    print(f"\n{'─' * 55}")
    print(f"Done. {saved} domain report(s) generated.")
    if args.source == "onedrive" and not args.no_upload:
        print("OneDrive will sync them to Teams in the background.")


if __name__ == "__main__":
    main()
