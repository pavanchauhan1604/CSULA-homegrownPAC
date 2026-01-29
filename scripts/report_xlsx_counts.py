import argparse
import pathlib
import sys

import openpyxl

# Ensure repo root (parent of scripts/) is on sys.path so `import config` works.
REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import config


def extract_url_from_hyperlink_formula(value: object) -> object:
    if isinstance(value, str) and value.startswith("=HYPERLINK("):
        # Typical form: =HYPERLINK("url","label")
        parts = value.split('"')
        if len(parts) >= 2:
            return parts[1]
    return value


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Print counts from the latest Excel report for a site."
    )
    parser.add_argument("--site", required=True, help="Site name, e.g. calstatela.edu_accessibility")
    args = parser.parse_args()

    xlsx_path = pathlib.Path(config.get_excel_report_path(args.site, prefer_latest=True))

    wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
    scanned_ws = wb["Scanned PDFs"]
    unique_ws = wb["Unique PDFs"]

    # Column A is PDF Title; column B is the PDF URL (hyperlink formula)
    scanned_values = [
        v
        for (v,) in scanned_ws.iter_rows(
            min_row=2, min_col=2, max_col=2, values_only=True
        )
        if v
    ]

    scanned_rows = len(scanned_values)
    unique_from_scanned = len({extract_url_from_hyperlink_formula(v) for v in scanned_values})

    # Column A is PDF Title; column B is the PDF URL
    unique_sheet_rows = sum(
        1
        for (v,) in unique_ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True)
        if v
    )

    print(f"xlsx_path: {xlsx_path}")
    print(f"scanned_rows: {scanned_rows}")
    print(f"unique_pdfs_from_scanned: {unique_from_scanned}")
    print(f"unique_sheet_rows: {unique_sheet_rows}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
