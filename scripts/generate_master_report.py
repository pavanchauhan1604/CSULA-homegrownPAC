"""Generate a historical master Excel report summarising all synced domains.

For every domain folder found in TEAMS_ONEDRIVE_PATH this script:
    1. Finds the latest .xlsx report inside that folder
    2. Reads the 'Unique PDFs' sheet
    3. Counts total unique PDFs and high-priority PDFs
    4. Appends one run snapshot to a historical Data sheet
    5. Refreshes a Dashboard sheet with a run-date dropdown selector

Usage
-----
        python scripts/generate_master_report.py
"""

from __future__ import annotations

import sys
import time
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

import config
from src.data_management.report_reader import (
    parse_domain_excel,
    find_latest_xlsx,
    folder_to_display_name,
    _TIMESTAMP_RE as _TS_RE,
)

SHEET_DATA = "Data"
SHEET_DASHBOARD = "Dashboard"
SHEET_RUN_INDEX = "Run Index"

HEADER_FILL = PatternFill("solid", fgColor="003262")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LOCK_RETRY_ATTEMPTS = 3
LOCK_RETRY_DELAY_SECONDS = 2


def _style_header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _ensure_workbook(path: Path) -> openpyxl.Workbook:
    if not path.exists():
        wb = openpyxl.Workbook()
    else:
        last_error: Exception | None = None
        wb = None
        for attempt in range(1, LOCK_RETRY_ATTEMPTS + 1):
            try:
                wb = openpyxl.load_workbook(path)
                break
            except PermissionError as e:
                last_error = e
                if attempt < LOCK_RETRY_ATTEMPTS:
                    time.sleep(LOCK_RETRY_DELAY_SECONDS)
        if wb is None:
            raise RuntimeError(
                f"Cannot open '{path}'. The file appears to be locked. "
                "Close Master/Master Report.xlsx in Excel (and let OneDrive finish syncing)"
                "then run this script again."
            ) from last_error

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]
    return wb


def _ensure_data_sheet(wb: openpyxl.Workbook):
    ws = wb[SHEET_DATA] if SHEET_DATA in wb.sheetnames else wb.create_sheet(SHEET_DATA, 0)
    _style_header_row(ws, ["Scan Date", "Domain", "Total Unique PDFs", "High Priority PDFs"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A2"
    
    # Add auto-filter arrows to the header row
    ws.auto_filter.ref = "A1:D1"
    
    return ws


def _ensure_run_index_sheet(wb: openpyxl.Workbook):
    ws = wb[SHEET_RUN_INDEX] if SHEET_RUN_INDEX in wb.sheetnames else wb.create_sheet(SHEET_RUN_INDEX)
    ws.sheet_state = "hidden"
    return ws


def _refresh_run_index(run_index_ws, data_ws) -> list[str]:
    run_values = {
        str(data_ws.cell(row=row, column=1).value).strip()
        for row in range(2, data_ws.max_row + 1)
        if data_ws.cell(row=row, column=1).value
    }
    ordered = sorted(run_values, reverse=True)

    run_index_ws.delete_rows(1, run_index_ws.max_row)
    for i, val in enumerate(ordered, start=1):
        run_index_ws.cell(row=i, column=1, value=val)
    run_index_ws.sheet_state = "hidden"
    return ordered


def _load_scan_timing() -> dict | None:
    """Read temp/scan_timing.json written by master_functions.create_all_pdf_reports()."""
    timing_path = config.TEMP_DIR / "scan_timing.json"
    if not timing_path.exists():
        return None
    try:
        import json
        with open(timing_path) as f:
            return json.load(f)
    except Exception:
        return None


def _refresh_dashboard(wb: openpyxl.Workbook, run_values: list[str], current_rows: list[tuple[str, str, int, int]]):
    ws = wb[SHEET_DASHBOARD] if SHEET_DASHBOARD in wb.sheetnames else wb.create_sheet(SHEET_DASHBOARD)

    # Clear all existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    ws["A1"] = "Master Report Dashboard"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Latest Scan Date"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = run_values[0] if run_values else ""

    # Dropdown for reference – shows all available scan dates in the hidden Run Index sheet.
    # NOTE: The domain table below always reflects the most recent roll-up (written as static values).
    # To view older runs, filter the Data sheet by Scan Date.
    list_end = max(1, len(run_values))
    dv = DataValidation(type="list", formula1=f"='{SHEET_RUN_INDEX}'!$A$1:$A${list_end}", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B3"])

    ws["A5"] = "Total Unique PDFs (latest roll-up)"
    ws["B5"] = sum(r[2] for r in current_rows)
    ws["B5"].alignment = Alignment(horizontal="center")
    ws["A6"] = "High Priority PDFs (latest roll-up)"
    ws["B6"] = sum(r[3] for r in current_rows)
    ws["B6"].alignment = Alignment(horizontal="center")

    # Scan timing — written by master_functions.create_all_pdf_reports()
    timing = _load_scan_timing()
    if timing:
        ws["A7"] = "Scan Started"
        ws["A7"].font = Font(bold=True)
        ws["B7"] = timing.get("scan_start", "")
        ws["A8"] = "Scan Completed"
        ws["A8"].font = Font(bold=True)
        ws["B8"] = timing.get("scan_end", "")
        ws["A9"] = "Scan Duration"
        ws["A9"].font = Font(bold=True)
        ws["B9"] = timing.get("duration_human", "")
        ws["B9"].font = Font(bold=True, color="003262")
        tip_row = 11
    else:
        tip_row = 8

    ws[f"A{tip_row}"] = "Tip: For older runs, go to the Data sheet and use Excel's column filter on 'Scan Date'."
    ws[f"A{tip_row}"].font = Font(italic=True, color="666666")

    table_header_row = tip_row + 2
    data_start_row = table_header_row + 1
    _style_header_row(ws, ["Domain", "Total Unique PDFs", "High Priority PDFs"], row=table_header_row)

    for i, (_, domain, total, high) in enumerate(current_rows, start=data_start_row):
        ws.cell(row=i, column=1, value=domain)
        ws.cell(row=i, column=2, value=total).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=high).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = f"A{data_start_row}"




# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    onedrive_path = Path(config.TEAMS_ONEDRIVE_PATH)
    if not onedrive_path.exists():
        print(f"[ERROR] OneDrive folder not found: {onedrive_path}")
        sys.exit(1)

    master_dir = onedrive_path / "Master"
    master_dir.mkdir(exist_ok=True)
    output_path = master_dir / "Master Report.xlsx"

    rows: list[tuple[str, str, int, int]] = []  # (scan_date, domain_display, total_unique, high_count)

    for folder in sorted(onedrive_path.iterdir()):
        if not folder.is_dir():
            continue
        # Skip special / meta folders
        if folder.name.startswith(".") or folder.name in {"Mail Drafts"}:
            continue

        xlsx = find_latest_xlsx(folder)
        if not xlsx:
            print(f"  [SKIP] No Excel found in: {folder.name}")
            continue

        metrics = parse_domain_excel(xlsx)
        if not metrics:
            print(f"  [SKIP] Could not read {xlsx.name}")
            continue

        total, high = metrics["unique_pdfs"], metrics["high_priority"]
        display = folder_to_display_name(folder.name)
        
        # Extract the logical scan date (YYYY-MM-DD) from the xlsx filename so that
        # the dashboard reflects the actual time the domain was scanned, rather than
        # the time this master script ran.
        m = _TS_RE.search(xlsx.name)
        scan_date = m.group(1)[:10] if m else datetime.today().strftime("%Y-%m-%d")
        
        rows.append((scan_date, display, total, high))
        print(f"  [OK] {display}: {total} unique PDFs, {high} high priority (scan: {scan_date})")

    # Sort by high priority PDFs descending
    rows.sort(key=lambda x: x[3], reverse=True)

    wb = _ensure_workbook(output_path)
    data_ws = _ensure_data_sheet(wb)
    run_index_ws = _ensure_run_index_sheet(wb)

    # Build set of already-recorded (scan_date, domain) pairs so re-runs
    # on the same day don't append duplicate rows.
    existing = {
        (str(data_ws.cell(row=r, column=1).value).strip(),
         str(data_ws.cell(row=r, column=2).value).strip())
        for r in range(2, data_ws.max_row + 1)
        if data_ws.cell(row=r, column=1).value
    }

    for scan_date, domain, total, high in rows:
        if (scan_date, domain) in existing:
            print(f"  [SKIP] Already recorded {domain} for {scan_date}")
            continue
        data_ws.append([scan_date, domain, total, high])
        data_ws.cell(row=data_ws.max_row, column=3).alignment = Alignment(horizontal="center")
        data_ws.cell(row=data_ws.max_row, column=4).alignment = Alignment(horizontal="center")

    run_values = _refresh_run_index(run_index_ws, data_ws)
    _refresh_dashboard(wb, run_values, rows)

    try:
        wb.save(output_path)
    except PermissionError as e:
        raise RuntimeError(
            f"Cannot save '{output_path}'. The file appears to be locked. "
            "Close Master/Master Report.xlsx in Excel (and let OneDrive finish syncing)"
            "then run this script again."
        ) from e

    print(f"\n[DONE] Master Report saved → {output_path}")
    print(f"       {len(rows)} domains | {sum(r[2] for r in rows)} total unique PDFs | {sum(r[3] for r in rows)} high priority")
    print(f"       {len(run_values)} scan dates available in Dashboard dropdown")


if __name__ == "__main__":
    main()
