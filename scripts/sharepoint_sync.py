"""Sync Excel reports to OneDrive/SharePoint and generate HTML email drafts.

Runs in two phases:

  Phase 1 — Sync (optional):
    For each domain in config.DOMAINS (or --domains filter), if a newer local
    Excel exists it is copied to the corresponding OneDrive folder.
    If there is no local Excel the domain is skipped with [NO LOCAL].

  Phase 2 — Draft generation:
    Iterates over EVERY folder already present in the OneDrive directory.
    Reads whatever Excel is already there (just synced or from a previous run)
    and writes/overwrites one HTML email draft per assigned employee.

This means that on a fresh start (no local Excels yet) Phase 1 produces nothing
to sync but Phase 2 still regenerates all drafts from the existing OneDrive files.

The send_emails.py script reads these drafts to send via Outlook COM.
No database access is required by either script.

Usage
-----
    # Full run (sync + regenerate all drafts):
    python scripts/sharepoint_sync.py

    # Limit sync phase to specific domains (draft phase still covers all folders):
    python scripts/sharepoint_sync.py --domains www.calstatela.edu_admissions
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl

import config
from src.core.filters import get_priority_level
from src.communication.communications import (
    template_email,
    create_html_email_summary,
    _display_domain,
)


# ---------------------------------------------------------------------------
# CSV loaders
# ---------------------------------------------------------------------------

def load_employees(csv_path: Path) -> dict:
    """Returns {employee_id: {first_name, last_name, email}}.

    employees.csv format (with header row):
        Full Name, Employee ID, Email
    """
    employees = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header
        for row in reader:
            if not row or not row[1].strip():
                continue
            parts = row[0].strip().split(" ", 1)
            employees[row[1].strip()] = {
                "first_name": parts[0],
                "last_name": parts[1] if len(parts) > 1 else "",
                "email": row[2].strip() if len(row) > 2 else "",
            }
    return employees


def load_site_assignments(csv_path: Path) -> dict:
    """Returns {security_group: [employee_id, ...]}.

    site_assignments.csv format (no header row):
        Security Group, Employee Name, Employee ID, Employee Email
    """
    assignments: dict = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or not row[0].strip():
                continue
            group = row[0].strip()
            employee_id = row[2].strip() if len(row) > 2 else ""
            if employee_id:
                assignments.setdefault(group, []).append(employee_id)
    return assignments


def load_domain_security_groups(csv_path: Path) -> dict:
    """Returns {domain_key: security_group_name}.

    sites.csv format (no header row):
        URL, Security Group
    """
    result = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or not row[0].strip():
                continue
            domain_key = config._url_to_domain_key(row[0])
            if len(row) > 1:
                group = row[1].strip()
                result[domain_key] = group
                # Also store without leading "www." so both forms match
                if domain_key.startswith("www."):
                    result[domain_key[4:]] = group
    return result


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def read_unique_pdfs_sheet(xlsx_path: Path) -> list:
    """Read 'Unique PDFs' sheet and return list of row dicts keyed by header."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Unique PDFs" not in wb.sheetnames:
        return []
    ws = wb["Unique PDFs"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(cell is not None for cell in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


# ---------------------------------------------------------------------------
# Priority derivation from Excel row
# ---------------------------------------------------------------------------

def _coerce_int(val, fallback: int = 0) -> int:
    try:
        if val is None or str(val).strip() in ("", "None"):
            return fallback
        return int(float(val))
    except (ValueError, TypeError):
        return fallback


def _coerce_bool(val) -> bool:
    """Convert Excel cell value to bool. Handles 0/1, True/False, Yes/No."""
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("1", "true", "yes", "1.0")


def row_to_priority_data(row: dict) -> dict:
    """Build the dict expected by filters.get_priority_level() from an Excel row."""
    tagged = 1 if _coerce_bool(row.get("tagged")) else 0

    # failed_checks may have been converted to Yes/No by a known export quirk.
    # Fall back to using the pre-computed Errors/Page column × page_count.
    failed_raw = row.get("failed_checks")
    if str(failed_raw).strip().lower() in ("yes", "no", ""):
        # Reconstruct from Errors/Page * page_count as best approximation.
        errors_per_page = _coerce_int(row.get("Errors/Page", 0))
        page_count = _coerce_int(row.get("page_count", 0))
        failed_checks = errors_per_page * page_count if page_count > 0 else 0
    else:
        failed_checks = _coerce_int(failed_raw)

    return {
        "tagged": tagged,
        "pdf_text_type": str(row.get("pdf_text_type") or ""),
        "violations": _coerce_int(row.get("violations", 0)),
        "failed_checks": failed_checks,
        "page_count": _coerce_int(row.get("page_count", 0)),
        "has_form": 1 if _coerce_bool(row.get("has_form")) else 0,
        "approved_pdf_exporter": _coerce_bool(row.get("approved_pdf_exporter")),
    }


def _strip_hyperlink(val) -> str:
    """Extract bare URL from an Excel HYPERLINK formula string or return as-is."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.startswith("=HYPERLINK("):
        import re
        m = re.search(r'=HYPERLINK\("([^"]+)"', s)
        return m.group(1) if m else s
    return s


# ---------------------------------------------------------------------------
# Email content builder
# ---------------------------------------------------------------------------

def build_domain_data_for_email(domain: str, pdf_rows: list) -> dict:
    """Build the data structure expected by create_html_email_summary()."""
    pdfs = []
    for row in pdf_rows:
        priority_data = row_to_priority_data(row)
        priority_level, color_code, priority_name = get_priority_level(priority_data)

        pdf_uri = _strip_hyperlink(row.get("pdf_uri") or row.get("pdf_title") or "")
        parent_uri = _strip_hyperlink(row.get("parent_uri") or "")

        pdfs.append({
            "filename": pdf_uri.split("/")[-1] if pdf_uri else "Unknown",
            "pdf_uri": pdf_uri,
            "violations": priority_data["violations"],
            "failed_checks": priority_data["failed_checks"],
            "page_count": priority_data["page_count"],
            "priority_level": priority_level,
            "priority_name": priority_name,
            "color_code": color_code,
            "parent_uri": parent_uri,
        })

    priority_order = {"high": 0, "medium": 1, "low": 2}
    pdfs.sort(key=lambda x: (priority_order[x["priority_level"]], -x["violations"]))

    return {domain: {"pdfs": pdfs, "box_folder": None}}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

import re as _re
from datetime import datetime as _dt

_TS_RE = _re.compile(r"(\d{4}-\d{2}-\d{2})_\d{2}-\d{2}-\d{2}\.xlsx$", _re.IGNORECASE)
_TS_SORT_RE = _re.compile(r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.xlsx$", _re.IGNORECASE)


def _find_latest_xlsx(folder: Path) -> Path | None:
    """Return the most recent .xlsx in *folder* sorted by filename timestamp."""
    def _key(p: Path):
        m = _TS_SORT_RE.search(p.name)
        return m.group(1) if m else ""
    candidates = [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")]
    return max(candidates, key=_key) if candidates else None


def _xlsx_report_date(xlsx_path: Path) -> str:
    """Extract a human-readable date from an Excel filename timestamp."""
    m = _TS_RE.search(xlsx_path.name)
    if m:
        return _dt.strptime(m.group(1), "%Y-%m-%d").strftime("%B %d, %Y")
    return xlsx_path.stem


def _folder_to_domain_key(folder_name: str) -> str:
    """Convert an OneDrive folder name back to a domain key.

    e.g. 'www-calstatela-edu_admissions' -> 'www.calstatela.edu_admissions'
    Only the part before the first '_' has dashes converted to dots.
    """
    if "_" in folder_name:
        domain_part, rest = folder_name.split("_", 1)
        return f"{domain_part.replace('-', '.')}_{rest}"
    return folder_name.replace("-", ".")


# ---------------------------------------------------------------------------
# Phase 1 — Sync local Excel → OneDrive
# ---------------------------------------------------------------------------

def sync_excel_to_onedrive(domain: str, onedrive_path: Path) -> bool:
    """Copy the latest local Excel for *domain* to its OneDrive folder.

    Returns True if a file was copied, False if no local Excel was found.
    """
    xlsx_path = config.get_excel_report_path(domain, prefer_latest=True)
    if not xlsx_path or not Path(xlsx_path).exists():
        print(f"  [NO LOCAL] {domain}")
        return False

    folder_name = config.get_domain_folder_name(domain)
    dest_folder = onedrive_path / folder_name
    dest_folder.mkdir(parents=True, exist_ok=True)

    dest_xlsx = dest_folder / xlsx_path.name
    shutil.copy2(xlsx_path, dest_xlsx)
    print(f"  [SYNC] {xlsx_path.name} → .../{folder_name}/")
    return True


# ---------------------------------------------------------------------------
# Phase 2 — Generate drafts from OneDrive folder
# ---------------------------------------------------------------------------

def generate_drafts_for_folder(
    folder: Path,
    employees: dict,
    assignments: dict,
    domain_groups: dict,
) -> bool:
    """Generate email drafts for a single OneDrive domain folder.

    Reads the latest Excel already present in the folder (just synced or from a
    previous run) and writes one HTML draft per assigned employee.
    Returns True if at least one draft was written.
    """
    xlsx_path = _find_latest_xlsx(folder)
    if not xlsx_path:
        print(f"  [SKIP] No Excel in OneDrive folder: {folder.name}")
        return False

    pdf_rows = read_unique_pdfs_sheet(xlsx_path)
    if not pdf_rows:
        print(f"  [WARN] 'Unique PDFs' sheet is empty — no drafts generated")
        return True

    # Resolve security group: prefer sites.csv, fall back to Excel column
    domain_key = _folder_to_domain_key(folder.name)
    security_group = domain_groups.get(domain_key, "")
    if not security_group:
        security_group = str(pdf_rows[0].get("security_group_name") or "").strip()

    employee_ids = assignments.get(security_group, [])
    if not employee_ids:
        print(f"  [WARN] No employees for security group '{security_group}' — no drafts generated")
        return True

    domain_data = build_domain_data_for_email(domain_key, pdf_rows)
    html_summary = create_html_email_summary(domain_data)
    display = _display_domain(domain_key)
    report_date = _xlsx_report_date(xlsx_path)

    drafts_folder = folder / "Mail Drafts"
    drafts_folder.mkdir(exist_ok=True)
    drafts_written = 0

    for employee_id in employee_ids:
        employee = employees.get(employee_id)
        if not employee:
            print(f"  [WARN] Employee ID '{employee_id}' not in employees.csv — skipped")
            continue

        email_html = template_email({
            "employee_first_name": employee["first_name"],
            "employee_full_name": f"{employee['first_name']} {employee['last_name']}".strip(),
            "pdf_data_table": html_summary,
            "report_date": report_date,
        })

        draft_path = drafts_folder / f"{employee_id}_draft.html"
        draft_path.write_text(email_html, encoding="utf-8")
        drafts_written += 1

    print(f"  [OK] {drafts_written} draft(s) → .../{folder.name}/Mail Drafts/ ({display})")
    return drafts_written > 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _check_onedrive() -> Path:
    raw = getattr(config, "TEAMS_ONEDRIVE_PATH", "")
    if not raw:
        print("ERROR: TEAMS_ONEDRIVE_PATH is not set in config.py")
        sys.exit(1)
    p = Path(raw)
    if not p.exists():
        print(f"ERROR: OneDrive path does not exist: {p}")
        print("Make sure your Teams channel Files folder is synced via OneDrive.")
        sys.exit(1)
    return p


def main() -> None:
    onedrive_path = _check_onedrive()

    parser = argparse.ArgumentParser(
        description="Sync Excel reports to OneDrive and generate HTML email drafts.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--domains",
        nargs="+",
        metavar="DOMAIN",
        help="Limit the sync phase to these domain(s). Draft generation always covers all OneDrive folders.",
    )
    args = parser.parse_args()

    sync_domains = args.domains or config.DOMAINS

    # Load CSVs once
    employees = load_employees(config.EMPLOYEES_CSV)
    assignments = load_site_assignments(config.SITE_ASSIGNMENTS_CSV)
    domain_groups = load_domain_security_groups(config.SITES_CSV)

    print(f"OneDrive folder : {onedrive_path}")
    print(f"Employees loaded: {len(employees)}")
    print(f"Assignments     : {sum(len(v) for v in assignments.values())} total across {len(assignments)} group(s)")

    # ── Phase 1: Sync local Excels → OneDrive ────────────────────────────────
    print(f"\n── Phase 1: Sync ({len(sync_domains)} domain(s)) ──")
    synced = 0
    for domain in sync_domains:
        if sync_excel_to_onedrive(domain, onedrive_path):
            synced += 1
    print(f"Synced {synced} / {len(sync_domains)} domain(s).\n")

    # ── Phase 2: Generate drafts for every OneDrive folder ───────────────────
    onedrive_folders = sorted(
        f for f in onedrive_path.iterdir()
        if f.is_dir() and not f.name.startswith(".")
    )
    print(f"── Phase 2: Generate drafts ({len(onedrive_folders)} folder(s) in OneDrive) ──")
    drafts_ok = drafts_skipped = 0
    for folder in onedrive_folders:
        print(f"▶ {folder.name}")
        try:
            if generate_drafts_for_folder(folder, employees, assignments, domain_groups):
                drafts_ok += 1
            else:
                drafts_skipped += 1
        except Exception as exc:
            print(f"  [ERROR] {exc}")
            drafts_skipped += 1
        print()

    print("─" * 55)
    print(f"Done.  Synced: {synced}  |  Drafts written: {drafts_ok}  |  Skipped: {drafts_skipped}")
    print("OneDrive will sync new files to Teams in the background.")


if __name__ == "__main__":
    main()
