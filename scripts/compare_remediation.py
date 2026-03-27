#!/usr/bin/env python3
"""
Standalone PDF accessibility remediation comparison tool.

Scans two folders of PDFs (originals and remediated versions) using
VeraPDF + pikepdf, matches pairs by the 6-char fingerprint prefix
in each filename, and outputs a side-by-side Excel comparison report.

Usage
-----
    python scripts/compare_remediation.py \\
        --originals "/path/to/originals" \\
        --remediated "/path/to/remediated" \\
        --output "/path/to/report.xlsx"          # optional, defaults to parent folder

No database access. No spider involvement. Completely standalone.
"""

from __future__ import annotations

import argparse
import os
import subprocess
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from src.core.pdf_priority import pdf_check, violation_counter

# ---------------------------------------------------------------------------
# Colours (CSULA brand)
# ---------------------------------------------------------------------------
NAVY        = "003262"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED   = "FCE4D6"
LIGHT_GREY  = "F2F2F2"
YELLOW_FILL = "FFF2CC"
GREEN_FONT  = "006400"
RED_FONT    = "8B0000"

# One temp JSON file per thread — avoids collisions between parallel workers.
_thread_local = threading.local()


def _temp_json() -> Path:
    tid = threading.get_ident()
    return config.TEMP_DIR / f"compare_{os.getpid()}_{tid}.json"


# ---------------------------------------------------------------------------
# Core scanner — works on a local file, no download needed
# ---------------------------------------------------------------------------

def scan_pdf_local(pdf_path: Path) -> dict:
    """Run VeraPDF + pikepdf on a local PDF. Returns a metrics dict."""
    result = {
        "path": str(pdf_path),
        "name": pdf_path.name,
        "violations": 0,
        "failed_checks": 0,
        "tagged": False,
        "pdf_text_type": "Unknown",
        "pages": 0,
        "has_form": False,
        "errors_per_page": 0.0,
        "error": None,
    }

    temp_json = _temp_json()
    try:
        # ── VeraPDF ─────────────────────────────────────────────────────────
        cmd = f'"{config.VERAPDF_COMMAND}" -f ua1 --format json "{pdf_path}" > "{temp_json}"'
        subprocess.run(cmd, shell=True, capture_output=True, text=True)

        if temp_json.exists():
            viol = violation_counter(str(temp_json))
            result["violations"]    = viol.get("violations", 0)
            result["failed_checks"] = viol.get("failed_checks", 0)

        # ── pikepdf ─────────────────────────────────────────────────────────
        meta = pdf_check(str(pdf_path))
        if meta and not meta.get("pdf_check_error"):
            result["tagged"]        = bool(meta.get("tagged", False))
            result["pdf_text_type"] = meta.get("pdf_text_type", "Unknown") or "Unknown"
            result["has_form"]      = bool(meta.get("has_form", False))
            pages = (meta.get("doc_data") or {}).get("pages", 0) or 0
            result["pages"] = pages
            if pages > 0:
                result["errors_per_page"] = round(result["failed_checks"] / pages, 2)
        elif meta and meta.get("pdf_check_error"):
            result["error"] = meta["pdf_check_error"]

    except Exception as exc:
        result["error"] = str(exc)
    finally:
        try:
            temp_json.unlink(missing_ok=True)
        except Exception:
            pass

    return result


# ---------------------------------------------------------------------------
# Fingerprint extraction
# ---------------------------------------------------------------------------

def _fingerprint(filename: str) -> str:
    """Return the 6-char prefix before '__', or the first 6 chars."""
    if "__" in filename:
        return filename.split("__")[0]
    return filename[:6]


# ---------------------------------------------------------------------------
# Folder scanner
# ---------------------------------------------------------------------------

def scan_folder(folder: Path, label: str, max_workers: int = 8) -> dict[str, dict]:
    """Scan all PDFs in *folder*. Returns {fingerprint: metrics}."""
    pdfs = sorted(p for p in folder.glob("*.pdf"))
    if not pdfs:
        print(f"  [WARN] No PDFs found in {folder}")
        return {}

    print(f"\nScanning {len(pdfs)} PDFs [{label}] with {max_workers} workers...")
    results: dict[str, dict] = {}

    def _task(pdf_path: Path):
        print(f"  [{label}] {pdf_path.name}")
        return _fingerprint(pdf_path.stem), scan_pdf_local(pdf_path)

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_task, p): p for p in pdfs}
        for fut in as_completed(futures):
            fp, metrics = fut.result()
            results[fp] = metrics

    return results


# ---------------------------------------------------------------------------
# Pair matching + status
# ---------------------------------------------------------------------------

def _status(orig: dict, remed: dict) -> str:
    delta_v = remed["violations"] - orig["violations"]
    if delta_v < 0:
        return "Improved"
    if delta_v == 0:
        # Tagged flip counts as improvement even if violation count unchanged
        if not orig["tagged"] and remed["tagged"]:
            return "Improved"
        return "No Change"
    return "Regressed"


def build_pairs(orig_results: dict, remed_results: dict) -> list[dict]:
    pairs = []
    for fp in sorted(set(orig_results) | set(remed_results)):
        has_orig  = fp in orig_results
        has_remed = fp in remed_results
        if has_orig and has_remed:
            orig  = orig_results[fp]
            remed = remed_results[fp]
            pairs.append({
                "status": _status(orig, remed),
                "fingerprint": fp,
                "orig":  orig,
                "remed": remed,
            })
        elif has_orig:
            pairs.append({"status": "Unpaired", "fingerprint": fp,
                          "orig": orig_results[fp], "remed": None})
        else:
            pairs.append({"status": "Unpaired (remediated only)", "fingerprint": fp,
                          "orig": remed_results[fp], "remed": None})
    return pairs


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor=NAVY)
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _hdr(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def build_excel(pairs: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_s = wb.active
    ws_s.title = "Summary"

    improved  = [p for p in pairs if p["status"] == "Improved"]
    no_change = [p for p in pairs if p["status"] == "No Change"]
    regressed = [p for p in pairs if p["status"] == "Regressed"]
    unpaired  = [p for p in pairs if "Unpaired" in p["status"]]
    paired    = [p for p in pairs if p["remed"] is not None]

    avg_viol_reduction = (
        sum(p["orig"]["violations"] - p["remed"]["violations"] for p in improved) / len(improved)
        if improved else 0
    )
    avg_epp_reduction = (
        sum(p["orig"]["errors_per_page"] - p["remed"]["errors_per_page"] for p in improved) / len(improved)
        if improved else 0
    )
    pct_improved = f"{len(improved)/len(paired)*100:.1f}%" if paired else "N/A"

    ws_s["A1"] = "PDF Accessibility Remediation Comparison"
    ws_s["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws_s["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_s["A2"].font = Font(italic=True, color="666666")

    summary_rows = [
        ("Paired PDFs scanned",                  len(paired)),
        ("Improved",                              len(improved)),
        ("No Change",                             len(no_change)),
        ("Regressed",                             len(regressed)),
        ("Unpaired (original only)",              len(unpaired)),
        (None, None),
        ("% PDFs Improved",                       pct_improved),
        ("Avg Violation Reduction (improved)",    f"{avg_viol_reduction:.1f}"),
        ("Avg Errors/Page Reduction (improved)",  f"{avg_epp_reduction:.2f}"),
    ]

    row = 4
    for label, value in summary_rows:
        if label is None:
            row += 1
            continue
        ws_s.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws_s.cell(row=row, column=2, value=value)
        cell.alignment = Alignment(horizontal="center")

        # Colour the key metric cells
        if label == "Improved":
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        elif label == "Regressed":
            cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
        row += 1

    ws_s.column_dimensions["A"].width = 45
    ws_s.column_dimensions["B"].width = 22

    # ── Comparison sheet ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Comparison")

    headers = [
        "File Name",
        "Status",
        "Violations\nBefore",
        "Violations\nAfter",
        "Δ Violations",
        "Errors/Page\nBefore",
        "Errors/Page\nAfter",
        "Δ Errors/Page",
        "Tagged\nBefore",
        "Tagged\nAfter",
        "Pages",
        "Text Type\nBefore",
        "Text Type\nAfter",
        "Has Form",
        "Notes",
    ]

    _hdr(ws, 1, headers)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    status_fill = {
        "Improved":                   PatternFill("solid", fgColor=LIGHT_GREEN),
        "No Change":                  PatternFill("solid", fgColor=YELLOW_FILL),
        "Regressed":                  PatternFill("solid", fgColor=LIGHT_RED),
        "Unpaired":                   PatternFill("solid", fgColor=LIGHT_GREY),
        "Unpaired (remediated only)": PatternFill("solid", fgColor=LIGHT_GREY),
    }

    # Sort: Regressed first (needs attention), then Improved, No Change, Unpaired
    sort_order = {"Regressed": 0, "Improved": 1, "No Change": 2,
                  "Unpaired": 3, "Unpaired (remediated only)": 4}
    sorted_pairs = sorted(pairs, key=lambda p: sort_order.get(p["status"], 99))

    CENTER = Alignment(horizontal="center")

    for row_idx, pair in enumerate(sorted_pairs, start=2):
        status = pair["status"]
        fill   = status_fill.get(status, PatternFill())
        orig   = pair["orig"]
        remed  = pair["remed"]

        if remed is None:
            # Unpaired — only original side
            clean_name = orig["name"]
            values = [
                clean_name, status,
                orig["violations"], "—", "—",
                orig["errors_per_page"], "—", "—",
                "Yes" if orig["tagged"] else "No", "—",
                orig["pages"],
                orig["pdf_text_type"], "—",
                "Yes" if orig["has_form"] else "No",
                orig.get("error") or "",
            ]
        else:
            delta_v   = remed["violations"]      - orig["violations"]
            delta_epp = round(remed["errors_per_page"] - orig["errors_per_page"], 2)
            # Strip _remediated suffix for cleaner display
            clean_name = orig["name"]
            values = [
                clean_name, status,
                orig["violations"],      remed["violations"],      delta_v,
                orig["errors_per_page"], remed["errors_per_page"], delta_epp,
                "Yes" if orig["tagged"]  else "No",
                "Yes" if remed["tagged"] else "No",
                orig["pages"],
                orig["pdf_text_type"],   remed["pdf_text_type"],
                "Yes" if orig["has_form"] else "No",
                (orig.get("error") or remed.get("error") or ""),
            ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

            # Centre numeric / short columns
            if col_idx in {3, 4, 5, 6, 7, 8, 9, 10, 11, 14}:
                cell.alignment = CENTER

            # Colour delta cells: green = better, red = worse
            if col_idx == 5 and isinstance(val, int):
                if val < 0:
                    cell.font = Font(color=GREEN_FONT, bold=True)
                elif val > 0:
                    cell.font = Font(color=RED_FONT, bold=True)
            if col_idx == 8 and isinstance(val, float):
                if val < 0:
                    cell.font = Font(color=GREEN_FONT, bold=True)
                elif val > 0:
                    cell.font = Font(color=RED_FONT, bold=True)

    col_widths = [50, 14, 12, 12, 13, 14, 13, 14, 10, 10, 8, 16, 16, 10, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n[DONE] Report saved → {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compare original vs remediated PDF accessibility results."
    )
    parser.add_argument("--originals",  required=True, help="Folder of original PDFs")
    parser.add_argument("--remediated", required=True, help="Folder of remediated PDFs")
    parser.add_argument("--output",     default=None,  help="Output .xlsx path (optional)")
    parser.add_argument("--workers",    type=int, default=8,
                        help="Max concurrent VeraPDF workers per folder (default: 8)")
    args = parser.parse_args()

    orig_folder  = Path(args.originals).expanduser().resolve()
    remed_folder = Path(args.remediated).expanduser().resolve()

    if not orig_folder.is_dir():
        print(f"ERROR: originals folder not found: {orig_folder}")
        sys.exit(1)
    if not remed_folder.is_dir():
        print(f"ERROR: remediated folder not found: {remed_folder}")
        sys.exit(1)

    output_path = (
        Path(args.output).expanduser()
        if args.output
        else orig_folder.parent / f"remediation_comparison_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    )

    config.TEMP_DIR.mkdir(parents=True, exist_ok=True)

    # Scan originals first, then remediated (each internally parallel)
    orig_results  = scan_folder(orig_folder,  "ORIGINAL",   max_workers=args.workers)
    remed_results = scan_folder(remed_folder, "REMEDIATED", max_workers=args.workers)

    pairs = build_pairs(orig_results, remed_results)

    # Console summary
    improved  = sum(1 for p in pairs if p["status"] == "Improved")
    no_change = sum(1 for p in pairs if p["status"] == "No Change")
    regressed = sum(1 for p in pairs if p["status"] == "Regressed")
    unpaired  = sum(1 for p in pairs if "Unpaired" in p["status"])
    paired    = len(pairs) - unpaired

    print(f"\n{'='*55}")
    print(f"  Total paired:  {paired}")
    print(f"  Improved:      {improved}  ({improved/paired*100:.1f}%)" if paired else "  Improved: 0")
    print(f"  No Change:     {no_change}")
    print(f"  Regressed:     {regressed}")
    print(f"  Unpaired:      {unpaired}")
    print(f"{'='*55}")

    build_excel(pairs, output_path)


if __name__ == "__main__":
    main()
